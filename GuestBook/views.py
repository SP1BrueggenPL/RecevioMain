from django.utils.translation import activate
from datetime import datetime
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from .forms import *
from django.contrib.auth import get_user_model
import base64
import openpyxl
from django.http import FileResponse
from .utils import generate_bhp_pdf
from django.contrib.auth.decorators import user_passes_test
from .task import close_expired_visitors_task
from django.contrib import auth
from .forms import UserLoginForm
from .utils import clean_next_url
from django.http import JsonResponse
# from .sms_gateway import SMSGateway  # Replaced by email notifications
from django.utils import timezone
from django.conf import settings
from django.http import HttpResponse
from django.db.models import Q
from django.db.models import Case, When, Value, BooleanField
from .models import TrustedVisitor
from django.utils.translation import gettext as _
from .models import Reservation, AdminProfile, ReservationCode
from .forms import ReservationForm
from .models import Company, Host
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Visitor
from django.utils.dateparse import parse_date
from django.db.models import Count
from django.utils import timezone
import json
from django.contrib.auth.decorators import login_required
import os, socket, random, string
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db.models import Q
from .models import Package
from .forms import PackageForm, ScanForm, LabelScanForm
import signal
from django.utils.translation import gettext as _
import threading
import time
from django.utils import timezone
from django.db import transaction
# --- FIRE & FORGET HELPERS (wrzuć w views.py, np. pod importami) ---
import time
from threading import Thread
from django.db import transaction

# --- FIRE & FORGET HELPERS ---
import time
from threading import Thread

def _mark_timeout_later(model_cls, pk, field_name, seconds=5):
    def _timer():
        time.sleep(seconds)
        obj = model_cls.objects.filter(pk=pk).only('id', field_name).first()
        if obj and getattr(obj, field_name) == 'pending':
            model_cls.objects.filter(pk=pk).update(**{field_name: 'timeout'})
    Thread(target=_timer, daemon=True).start()

def print_badge_async(visitor_id, zpl_payload, soft_timeout=5, printer_ip='10.30.40.150', printer_port=9100):
    from GuestBook.models import Visitor
    def _worker():
        try:
            send_zpl_to_printer(zpl_payload, printer_ip=printer_ip, port=printer_port)
            Visitor.objects.filter(id=visitor_id, print_status='pending').update(print_status='printed')
        except Exception:
            Visitor.objects.filter(id=visitor_id, print_status='pending').update(print_status='error')
    Thread(target=_worker, daemon=True).start()
    _mark_timeout_later(Visitor, visitor_id, 'print_status', seconds=soft_timeout)

# def send_sms_async(visitor_id, number, message, soft_timeout=5):  # SMS disabled
#     from GuestBook.models import Visitor
#     def _worker():
#         try:
#             sms = SMSGateway()
#             res = sms.send_sms(number, message)
#             ok = bool(res and res.get('status') == 'success')
#             Visitor.objects.filter(id=visitor_id, sms_status='pending').update(
#                 sms_status='sent' if ok else 'error'
#             )
#         except Exception:
#             Visitor.objects.filter(id=visitor_id, sms_status='pending').update(sms_status='error')
#     Thread(target=_worker, daemon=True).start()
#     _mark_timeout_later(Visitor, visitor_id, 'sms_status', seconds=soft_timeout)

def send_email_to_host(visitor_id, host_email, subject, message, soft_timeout=5):
    """Send email notification to host asynchronously (replaces SMS)."""
    from GuestBook.models import Visitor
    from django.core.mail import send_mail
    from django.conf import settings as _s

    def _worker():
        try:
            if not host_email:
                Visitor.objects.filter(id=visitor_id, sms_status='pending').update(sms_status='skipped')
                return
            send_mail(subject, message, _s.DEFAULT_FROM_EMAIL, [host_email], fail_silently=False)
            Visitor.objects.filter(id=visitor_id, sms_status='pending').update(sms_status='sent')
        except Exception:
            Visitor.objects.filter(id=visitor_id, sms_status='pending').update(sms_status='error')

    Thread(target=_worker, daemon=True).start()
    _mark_timeout_later(Visitor, visitor_id, 'sms_status', seconds=soft_timeout)


ALPHABET = string.digits + string.ascii_uppercase  # baza36



# Timeout helper
class TimeoutException(Exception):
    pass

def _timeout_handler(signum, frame):
    raise TimeoutException()

from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError

def _get_printer_for_user(user):
    """Returns (ip, port) from AdminProfile, or defaults."""
    if user and user.is_authenticated:
        profile = AdminProfile.objects.filter(user=user).first()
        if profile and profile.printer_address:
            return profile.printer_address, profile.printer_port
    return '10.30.40.150', 9100

def run_with_timeout(func, *args, seconds=5, **kwargs):
    """
    Uruchamia func(*args, **kwargs) z limitem czasu.
    Zwraca: 'ok' (bez wyjątku), 'timeout' (przekroczony limit), 'error' (wyjątek).
    Semantyka zgodna z Twoją obecną implementacją na SIGALRM.
    """
    with ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(func, *args, **kwargs)
        try:
            _ = fut.result(timeout=seconds)  # ignorujemy wartość, liczy się brak wyjątku
            return 'ok'
        except FuturesTimeoutError:
            return 'timeout'
        except Exception:
            return 'error'

def send_email_with_timeout(subject: str, body: str, to_list: list[str], seconds: int = 5):
    """
    Wysyła e-mail z limitem czasu. Zwraca: 'ok' / 'timeout' / 'error'.
    """
    def _do_send():
        EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, to_list).send()
    return run_with_timeout(_do_send, seconds=seconds)

def _is_yes(value) -> bool:
    return str(value or "").strip().lower() in {"tak", "yes", "true", "1"}

def _company_display(visitor) -> str:
    return visitor.factory.company_name if visitor.factory else (visitor.company_name_text or "Brak")

def generate_package_code(prefix="BX", length=8):
    body = "".join(random.choices(ALPHABET, k=length))
    return f"{prefix}{body}"

def _extract_visit_purpose(data):
    if data.get('visit_purpose_choice') == 'other':
        return data.get('other_purpose', '')
    return data.get('visit_purpose_choice', '') or ''

def render_box_label_from_file(code: str, sender: str, recipient: str, template_path=None):
    """
    Wczytuje Twój szablon ZPL i podstawia {code}, {sender}, {recipient}.
    Domyślna ścieżka: <BASE_DIR>/GuestBook/Print_templates/box_label.zpl
    """
    if not template_path:
        template_path = os.path.join(settings.BASE_DIR, "GuestBook", "Print_templates", "BoxLabelTemplate.zpl")
    with open(template_path, encoding="utf-8") as f:
        tpl = f.read()
    return tpl.format(code=code, sender=sender, recipient=recipient)

def send_zpl_to_printer(zpl_data: str, printer_ip='10.30.40.150', port=9100):
    """Wysyła surowy kod ZPL do drukarki Zebra przez TCP/IP (port 9100).
    Jedyna niezawodna metoda dla ZPL — bezpośrednie połączenie RAW.
    Ścieżki UNC (\\server\\printer) nie są obsługiwane — użyj adresu IP.
    """
    is_unc = printer_ip.startswith('\\\\') or printer_ip.startswith('//')
    if is_unc:
        raise ValueError(
            f"Ścieżka UNC '{printer_ip}' nie jest obsługiwana dla ZPL. "
            "Podaj bezpośredni adres IP drukarki Zebra (np. 10.30.40.150) i port 9100."
        )
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(10)
    sock.connect((printer_ip, port))
    sock.sendall(zpl_data.encode('utf-8'))
    sock.close()

try:
    close_expired_visitors_task(repeat=3600)
except Exception:
    pass  # DB not yet migrated (e.g. first deploy); background_task table created on migrate

def is_reception(user):
    return user.groups.filter(name="Reception").exists()

def is_admin(user):
    return user.groups.filter(name="Admin").exists()

def is_user(user):
    return user.groups.filter(name="User").exists()

#@user_passes_test(is_reception)
def anonymous_required(function=None, redirect_url=None):

   if not redirect_url:
       redirect_url = 'guestbook'

   actual_decorator = user_passes_test(
       lambda u: u.is_anonymous,
       login_url=redirect_url
   )

   if function:
       return actual_decorator(function)
   return actual_decorator

User = get_user_model()


def index_redirect(request):
    return redirect('login')


from django.contrib import messages
from django.utils.http import url_has_allowed_host_and_scheme

def clean_next_url(next_url: str) -> str:
    # Twoja implementacja – przykład kontroli bezpieczeństwa
    return next_url if url_has_allowed_host_and_scheme(next_url, allowed_hosts=None) else ''

def login(request):
    context = {}

    if request.method == 'GET':
        form = UserLoginForm()
        next_param = clean_next_url(request.GET.get('next', ''))
        context['form'] = form
        context['next'] = next_param
        return render(request, 'login.html', context)

    if request.method == 'POST':
        form = UserLoginForm(request.POST)
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = auth.authenticate(username=username, password=password)

        if user:
            auth.login(request, user)
            next_url = clean_next_url(request.POST.get('next', ''))

            if user.is_superuser:
                return redirect('/nimda/')
            elif next_url:
                return redirect(next_url)
            elif 'Recevio_Reception' in user.groups.values_list('name', flat=True):
                return redirect('statistics')
            elif 'Recevio_User' in user.groups.values_list('name', flat=True):
                return redirect('reservation')
            elif 'Recevio_Helpdesk' in user.groups.values_list('name', flat=True):
                return redirect('dashboard')
            else:
                return redirect('profile')
        else:
            messages.error(request, "Incorrect login or password.", extra_tags='login')

        return redirect('login')

    return render(request, 'login.html', context)


@login_required
def logout(request):
    auth.logout(request)
    return redirect('login')

def start_registration(request):
    return render(request, 'kiosk/start.html')

def choose_language(request):
    if request.method == 'POST':
        lang = request.POST.get('lang')
        request.session[settings.LANGUAGE_COOKIE_NAME] = lang
        activate(lang)
        return redirect('choose_method')
    return render(request, 'kiosk/language_selection.html')

def choose_method_view(request):
    return render(request, 'kiosk/choose_method.html')

def enter_visitor_id(request):
    if request.method == 'POST':
        visitor_id = request.POST.get('visitor_id')
        request.session['visitor_id'] = visitor_id
        request.session['registration_mode'] = 'manual'
        return redirect('guest_form', visitor_id=visitor_id)
    return render(request, 'kiosk/enter_id.html')

def guest_form(request, visitor_id):
    lang = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')
    activate(lang)
    request.session[settings.LANGUAGE_COOKIE_NAME] = lang
    request.LANGUAGE_CODE = lang

    if request.method == 'POST':
        form = VisitorForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data

            visit_purpose = _extract_visit_purpose(data)

            factory = data.get('factory')
            company_name_text = ""
            host = data.get('host')

            if not factory:
                if data.get('other_company'):
                    company_name_text = data['other_company'].strip()

            request.session['visitor_data'] = {
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'phone': data['phone'],
                'factory_id': factory.id if factory else None,
                'company_name_text': company_name_text,         # ⬅️ ważne
                'visit_purpose': visit_purpose,
                'host_id': host.id if host else None,
                'audit_start_date': data['audit_start_date'].isoformat() if data.get('audit_start_date') else '',
                'audit_end_date': data['audit_end_date'].isoformat() if data.get('audit_end_date') else '',
            }

            request.session['is_audit'] = data.get('visit_purpose_choice') == 'audit'
            return redirect('safety_pdf_only', visitor_id=visitor_id)
    else:
        form = VisitorForm()

    return render(request, 'kiosk/guest_form.html', {'form': form})


def get_host_for_company(request):
    company_id = request.GET.get('company_id')
    try:
        company = Company.objects.get(id=company_id)
        if company.host_name:
            return JsonResponse({'host_id': company.host_name.id, 'host_name': company.host_name.host_name})
    except Company.DoesNotExist:
        pass
    return JsonResponse({}, status=404)


def safety_pdf_only(request, visitor_id):
    if request.method == 'POST':
        return redirect('safety_question1', visitor_id=visitor_id)
    return render(request, 'kiosk/safety_image_only.html', {'visitor_id': visitor_id})


def safety_question1_view(request, visitor_id):
    visitor_data = request.session.get("visitor_data")
    if not visitor_data:
        return redirect("start")

    is_audit = request.session.get("is_audit")
    audit_start_str = request.session.get("audit_start")

    if is_audit and audit_start_str:
        from datetime import date
        audit_start = date.fromisoformat(audit_start_str)
        if date.today() != audit_start:
            return redirect("signature_form", visitor_id=visitor_id)

    if request.method == "POST":
        q1 = request.POST.get("question_1")

        # ✅ Jeśli odpowiedź jest TAK/YES → blokujemy
        if q1.lower() in ["tak", "yes"]:
            request.session.flush()  # czyścimy całą sesję
            return render(request, "kiosk/access_denied.html")

        request.session["safety_answers"] = {"question_1": q1}
        return redirect("safety_question2", visitor_id=visitor_id)

    return render(request, "kiosk/safety_question_1.html", {'visitor_id': visitor_id})

def safety_question2_view(request, visitor_id):
    if "safety_answers" not in request.session:
        return redirect("safety_question1", visitor_id=visitor_id)

    if request.method == "POST":
        q2 = request.POST.get("question_2")
        safety = request.session.get("safety_answers", {})
        safety["question_2"] = q2
        request.session["safety_answers"] = safety  # <- kluczowy krok
        return redirect("safety_question3", visitor_id=visitor_id)

    return render(request, "kiosk/safety_question_2.html", {'visitor_id': visitor_id})

def safety_question3_view(request, visitor_id):
    if "safety_answers" not in request.session:
        return redirect("safety_question1", visitor_id=visitor_id)

    if request.method == "POST":
        q3 = request.POST.get("question_3")
        safety = request.session.get("safety_answers", {})
        safety["question_3"] = q3
        request.session["safety_answers"] = safety  # <- kluczowy krok
        return redirect("prepare_visitor", visitor_id=visitor_id)

    return render(request, "kiosk/safety_question_3.html", {'visitor_id': visitor_id})

def prepare_visitor(request, visitor_id):
    data = request.session.get('visitor_data', {})
    language = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')
    safety_data = request.session.get("safety_answers", {})
    host_id = data.get('host_id')
    host = Host.objects.filter(id=host_id).first() if host_id else None
    q1 = safety_data.get("question_1")
    q2 = safety_data.get("question_2")
    q3 = safety_data.get("question_3")

    factory_id = data.get('factory_id')
    factory = Company.objects.filter(id=factory_id).first() if factory_id else None
    company_name_text = data.get('company_name_text', '')  # <-- bierzemy tekst z sesji

    if request.session.get('is_audit'):
        start = request.session.get('audit_start')
        skip_safety = start and str(date.today()) != start
    else:
        skip_safety = False

    audit_start = data.get("audit_start_date")
    audit_end = data.get("audit_end_date")

    visitor = Visitor.objects.filter(visitor_id=visitor_id, end_time__isnull=True).order_by('-start_time').first()

    if not visitor:
        visitor = Visitor(
            visitor_id=visitor_id,
            first_name=data.get('first_name'),
            last_name=data.get('last_name'),
            phone=data.get('phone'),
            factory=factory,  # FK jeśli wybrano z listy
            company_name_text=company_name_text,  # tekst jeśli wpisano ręcznie
            visit_purpose=data.get('visit_purpose'),
            host=host,
            production_area=True,
            safety_acknowledged=True,
            language=language,
        )
    else:
        # jeśli w międzyczasie zmieniły się dane w sesji, zaktualizuj też istniejącego
        visitor.factory = factory
        if company_name_text:  # nie nadpisuj pustką
            visitor.company_name_text = company_name_text

    visitor.audit_start_date = audit_start or None
    visitor.audit_end_date = audit_end or None
    visitor.with_supervision = True
    visitor.safety_question_1 = q1
    visitor.safety_question_2 = q2
    visitor.safety_question_3 = q3
    visitor.save()

    return redirect('signature_form', visitor_id=visitor_id)


def supervision_form(request, visitor_id):
    if request.method == 'POST':
        form = SupervisionForm(request.POST)
        if form.is_valid():
            data = request.session.get('visitor_data', {})
            language = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')
            safety_data = request.session.get("safety_answers", {})
            host_id = data.get('host_id')
            host = Host.objects.filter(id=host_id).first() if host_id else None
            q1 = safety_data.get("question_1")
            q2 = safety_data.get("question_2")
            q3 = safety_data.get("question_3")
            factory_id = data.get('factory_id')
            factory = Company.objects.filter(id=factory_id).first() if factory_id else None
            visitor = Visitor.objects.filter(visitor_id=visitor_id, end_time__isnull=True).order_by('-start_time').first()

            if not visitor:
                visitor = Visitor(
                    visitor_id=visitor_id,
                    first_name=data.get('first_name'),
                    last_name=data.get('last_name'),
                    phone=data.get('phone'),
                    factory=factory,
                    company_name_text=data.get('company_name_text', ''),  # <-- DODAJ
                    visit_purpose=data.get('visit_purpose'),
                    host=host,
                    production_area=request.session.get('production_area'),
                    safety_acknowledged=request.session.get('safety_acknowledged'),
                    language=language,
                )
            else:
                # aktualizacja na wszelki wypadek
                visitor.factory = factory
                if data.get('company_name_text'):
                    visitor.company_name_text = data.get('company_name_text')

            visitor.with_supervision = form.cleaned_data['with_supervision'] == 'True'
            visitor.safety_question_1 = q1
            visitor.safety_question_2 = q2
            visitor.safety_question_3 = q3
            visitor.save()

            return redirect('signature_form', visitor_id=visitor_id)
    else:
        form = SupervisionForm()
    return render(request, 'kiosk/supervision_form.html', {'form': form})


def signature_form(request, visitor_id):
    if request.method == 'POST':
        data_url = request.POST.get('signature_data')
        format, imgstr = data_url.split(';base64,')
        data = ContentFile(base64.b64decode(imgstr), name=f"{visitor_id}_signature.png")
        path = default_storage.save(f'signatures/{data.name}', data)

        # Zaktualizuj istniejącego Visitora
        visitor = Visitor.objects.filter(visitor_id=visitor_id, end_time__isnull=True).order_by('-start_time').first()
        visitor.with_supervision = request.session.get('with_supervision', True)
        if visitor:
            visitor.signed = path
            visitor.save()

        return redirect('finish_registration', visitor_id=visitor_id)

    return render(request, 'kiosk/signature_form.html')



def finish_registration(request, visitor_id):
    visitor = (Visitor.objects
               .filter(visitor_id=visitor_id, end_time__isnull=True)
               .order_by('-start_time').first())
    if not visitor:
        return render(request, 'kiosk/error.html', {'message': _("Visitor not found or already completed.")})

    if not visitor.language:
        visitor.language = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')

    mode = request.session.get('registration_mode')
    if mode == 'manual':
        visitor.approved = False
        visitor.known_guest = False
        visitor.safety_acknowledged = True
    elif mode == 'badge':
        visitor.approved = True
        visitor.known_guest = True
        visitor.safety_acknowledged = True
        badge_id = request.session.get('badge_id')
        if badge_id and badge_id == visitor.visitor_id:
            if not TrustedVisitor.objects.filter(badge_id=badge_id).exists():
                TrustedVisitor.objects.create(
                    first_name=visitor.first_name,
                    last_name=visitor.last_name,
                    phone=visitor.phone,
                    company=_company_display(visitor),
                    visit_purpose=visitor.visit_purpose,
                    host_name=visitor.host.host_name if visitor.host else "",
                    host_phone=visitor.host.phone if visitor.host else "",
                    language=visitor.language,
                    production_area=visitor.production_area,
                    with_supervision=bool(visitor.with_supervision),
                    safety_acknowledged=True,
                    badge_id=badge_id,
                )

    # --- Generowanie ZPL do druku po stronie klienta (Zebra Browser Print) ---
    zpl_for_print = None
    if visitor.production_area:
        try:
            template_path = os.path.join(settings.BASE_DIR, 'GuestBook', 'Print_templates', 'template_zebra.zpl')
            with open(template_path, encoding='utf-8') as tpl:
                zpl_template = tpl.read()
            zpl_for_print = zpl_template.format(
                company=_company_display(visitor),
                first_name=visitor.first_name,
                last_name=visitor.last_name,
                visit_purpose=visitor.visit_purpose or '',
                supervisor=visitor.host.host_name if visitor.host else ''
            )
            visitor.print_status = "pending"
        except Exception:
            visitor.print_status = "error"
    else:
        visitor.print_status = "skipped"
    visitor.save(update_fields=["print_status"])

    # --- Email do gospodarza fire-and-forget (SMS disabled) ---
    try:
        host_email = visitor.host.email if visitor.host else ''
        status_alergen = (
            "Uwaga - Gość jest uczulony na jeden lub więcej naszych alergenów."
            if _is_yes(visitor.safety_question_3) else ""
        )
        company_for_msg = _company_display(visitor)
        message = (
            f"Informacja o przybyciu gościa.\n\n"
            f"Dane:\n"
            f"Imię i nazwisko: {visitor.first_name} {visitor.last_name}\n"
            f"Telefon: {visitor.phone}\n"
            f"Firma: {company_for_msg}\n"
            f"Cel wizyty: {visitor.visit_purpose}\n\n"
            f"Gość oczekuje na odebranie z recepcji.\n\n"
            f"{status_alergen}"
        )
        visitor.sms_status = "pending"
        visitor.save(update_fields=["sms_status"])
        send_email_to_host(visitor.id, host_email, 'New visitor arrived', message)
    except Exception:
        visitor.sms_status = "error"
        visitor.save(update_fields=["sms_status"])

    # posprzątaj sesję i zapisz
    for key in ['badge_id', 'registration_mode', 'visitor_data', 'safety_answers']:
        request.session.pop(key, None)
    visitor.save()

    # kolor etykiety
    if visitor.production_area:
        badge_color = _("Red (moving around the factory only with a host)") if visitor.with_supervision \
                      else _("Green (moving around the factory without a host)")
        badge_css = "red" if visitor.with_supervision else "green"
    else:
        badge_color = badge_css = None

    return render(request, 'kiosk/complete.html', {
        'badge_color': badge_color, 'badge_css': badge_css,
        'is_production': visitor.production_area,
        'zpl_for_print': zpl_for_print,
        'visitor_id': visitor.pk,
    })



# dla pestki

def enter_badge_view(request):
    if request.method == 'POST':
        badge_id = request.POST.get('badge_id')
        request.session['registration_mode'] = 'badge'
        request.session['badge_id'] = badge_id

        trusted = TrustedVisitor.objects.filter(badge_id=badge_id).first()
        if trusted:
            # Firma z TrustedVisitor (tekst); próbujemy dopasować istniejący Company
            company_name_str = (trusted.company or "").strip()
            factory = Company.objects.filter(company_name__iexact=company_name_str).first() if company_name_str else None

            # Host po nazwie (jeśli jest)
            host = Host.objects.filter(host_name__iexact=trusted.host_name.strip()).first() if trusted.host_name else None

            # Dane do sesji na dalsze etapy
            request.session['visitor_data'] = {
                'first_name': trusted.first_name,
                'last_name': trusted.last_name,
                'phone': trusted.phone,
                'factory_id': factory.id if factory else None,   # FK jeśli istnieje
                'company_name_text': "" if factory else company_name_str,  # tekst gdy nie mamy FK
                'visit_purpose': trusted.visit_purpose or "",
                'host_id': host.id if host else None,
                'audit_start_date': "",
                'audit_end_date': "",
            }

            # Kopia „informacyjna" (opcjonalne)
            request.session['trusted_visitor'] = {
                'first_name': trusted.first_name,
                'last_name': trusted.last_name,
                'phone': trusted.phone,
                'company': company_name_str,
                'host_name': trusted.host_name or "",
                'host_phone': trusted.host_phone or "",
                'visit_purpose': trusted.visit_purpose or ""
            }

            return redirect('production_form', visitor_id=badge_id)

        # brak trusted → zwykła ścieżka z pestką
        return redirect('guest_form_badge')

    return render(request, 'kiosk/enter_badge.html')




def guest_form_badge(request):
    lang = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')
    activate(lang)
    request.session[settings.LANGUAGE_COOKIE_NAME] = lang
    request.LANGUAGE_CODE = lang

    badge_id = request.session.get('badge_id')
    if not badge_id:
        return redirect('enter_badge')

    visitor_id = badge_id

    if request.method == 'POST':
        form = VisitorForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            visit_purpose = _extract_visit_purpose(data)

            factory = data.get('factory')  # FK lub None
            company_name_text = ""
            host = data.get('host')

            if not factory and data.get('other_company'):
                company_name_text = data['other_company'].strip()

            request.session['visitor_data'] = {
                'first_name': data['first_name'],
                'last_name': data['last_name'],
                'phone': data['phone'],
                'factory_id': factory.id if factory else None,
                'company_name_text': company_name_text,
                'visit_purpose': visit_purpose,
                'host_id': host.id if host else None,
                'audit_start_date': data['audit_start_date'].isoformat() if data.get('audit_start_date') else '',
                'audit_end_date': data['audit_end_date'].isoformat() if data.get('audit_end_date') else '',
            }

            request.session['is_audit'] = data.get('visit_purpose_choice') == 'audit'
            return redirect('production_form', visitor_id=visitor_id)
    else:
        form = VisitorForm()

    return render(request, 'kiosk/guest_form.html', {'form': form, 'visitor_id': visitor_id})



def production_form(request, visitor_id):
    if request.method == 'POST':
        choice = request.POST.get('production_area')
        request.session['production_area'] = True if choice == 'True' else False

        return redirect('received_badge', visitor_id=visitor_id)

    return render(request, 'kiosk/production_form.html', {'visitor_id': visitor_id})


def received_badge_view(request, visitor_id):
    if request.method == 'POST':
        has_badge = request.POST.get('has_badge')
        production_area = request.session.get('production_area', False)

        # ⬇️ ZAPAMIĘTAJ W SESJI
        request.session['has_badge'] = (has_badge == 'True')

        if has_badge == 'True':
            return redirect('enter_visitor_id_for_badge', visitor_id=visitor_id)
        else:
            if production_area:
                return redirect('safety_pdf_trusted', visitor_id=visitor_id)
            else:
                return redirect('signature_trusted', visitor_id=visitor_id)

    return render(request, 'kiosk/received_badge.html', {'visitor_id': visitor_id})



def enter_visitor_id_for_badge(request, visitor_id):
    if request.method == 'POST':
        manual_id = request.POST.get('visitor_id')
        request.session['manual_visitor_id'] = manual_id
        request.session['visitor_id'] = manual_id
        return redirect('signature_trusted', visitor_id=visitor_id)
    return render(request, 'kiosk/enter_id_for_badge.html')


def safety_pdf_trusted_view(request, visitor_id):
    if request.method == 'POST':
        return redirect('bhp_question1_trusted', visitor_id=visitor_id)
    return render(request, 'kiosk/safety_image_only.html', {'visitor_id': visitor_id})


def bhp_question1_trusted_view(request, visitor_id):
    if request.method == 'POST':
        q1 = request.POST.get("question_1")

        if q1.lower() in ["tak", "yes"]:
            request.session.flush()
            return render(request, "kiosk/access_denied.html")

        request.session["safety_answers"] = {"question_1": q1}
        return redirect("bhp_question2_trusted", visitor_id=visitor_id)

    return render(request, "kiosk/bhp_question_1_trusted.html", {'visitor_id': visitor_id})


def bhp_question2_trusted_view(request, visitor_id):
    if request.method == 'POST':
        q2 = request.POST.get("question_2")
        safety_answers = request.session.get("safety_answers", {})
        safety_answers["question_2"] = q2
        request.session["safety_answers"] = safety_answers
        return redirect("bhp_question3_trusted", visitor_id=visitor_id)

    return render(request, "kiosk/bhp_question_2_trusted.html", {'visitor_id': visitor_id})


def bhp_question3_trusted_view(request, visitor_id):
    if request.method == 'POST':
        q3 = request.POST.get("question_3")
        safety_answers = request.session.get("safety_answers", {})
        safety_answers["question_3"] = q3
        request.session["safety_answers"] = safety_answers
        request.session['safety_acknowledged'] = True
        return redirect("signature_trusted", visitor_id=visitor_id)

    return render(request, "kiosk/bhp_question_3_trusted.html", {'visitor_id': visitor_id})


def signature_trusted_view(request, visitor_id):
    if request.method == 'POST':
        data_url = request.POST.get('signature_data')
        format, imgstr = data_url.split(';base64,')
        data = ContentFile(base64.b64decode(imgstr), name=f"{visitor_id}_signature.png")
        path = default_storage.save(f'signatures/{data.name}', data)

        # Pobierz istniejącego Visitora (jeśli jest)
        manual_id = request.session.get('manual_visitor_id')
        real_id = manual_id or visitor_id

        visitor = Visitor.objects.filter(visitor_id=real_id, end_time__isnull=True).order_by('-start_time').first()

        if not visitor:
            visitor_data = request.session.get("visitor_data")
            if visitor_data:
                visitor = Visitor.objects.create(
                    visitor_id=real_id,
                    first_name=visitor_data.get("first_name", ""),
                    last_name=visitor_data.get("last_name", ""),
                    phone=visitor_data.get("phone", ""),
                    factory=Company.objects.filter(id=visitor_data.get("factory_id")).first() if visitor_data.get(
                        "factory_id") else None,
                    company_name_text=visitor_data.get("company_name_text", ""),  # ⬅️ DODANE
                    visit_purpose=visitor_data.get("visit_purpose", ""),
                    host=Host.objects.filter(id=visitor_data.get("host_id")).first() if visitor_data.get(
                        "host_id") else None,
                    production_area=request.session.get("production_area", False),
                    with_supervision=True,
                    audit_start_date=visitor_data.get("audit_start_date") or None,
                    audit_end_date=visitor_data.get("audit_end_date") or None,
                    language=request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl'),
                )
                request.session['visitor_id'] = real_id
            else:
                messages.error(request, _("Visitor not found. Please start again."))
                return redirect('enter_badge')

        # Aktualizuj dane Visitora
        visitor.signed = path
        visitor.with_supervision = False
        visitor.safety_acknowledged = request.session.get('safety_acknowledged', False)
        visitor.language = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')

        safety_answers = request.session.get("safety_answers", {})
        visitor.safety_question_1 = safety_answers.get("question_1")
        visitor.safety_question_2 = safety_answers.get("question_2")
        visitor.safety_question_3 = safety_answers.get("question_3")

        visitor.approved = True
        visitor.known_guest = True  # zawsze True

        # ✅ Zamknięcie tylko dla pestek (>3 znaki), ID manualne zostaje otwarte do ręcznego zamknięcia
        if len(str(visitor.visitor_id)) > 3:
            # nic nie ustawiamy, auto-close zrobi task
            pass

        visitor.save()

        # Dodaj TrustedVisitor jeśli jeszcze nie istnieje
        if not TrustedVisitor.objects.filter(badge_id=visitor.visitor_id).exists():
            TrustedVisitor.objects.create(
                first_name=visitor.first_name,
                last_name=visitor.last_name,
                phone=visitor.phone or "",
                company=_company_display(visitor),  # ⬅️ zamiast tylko factory
                visit_purpose=visitor.visit_purpose,
                host_name=visitor.host.host_name if visitor.host else "",
                host_phone=visitor.host.phone if visitor.host else "",
                language=visitor.language,
                production_area=visitor.production_area,
                with_supervision=visitor.with_supervision or False,
                safety_acknowledged=visitor.safety_acknowledged,
                badge_id=visitor.visitor_id
            )

        for key in ['trusted_visitor', 'safety_answers', 'badge_id', 'registration_mode']:
            request.session.pop(key, None)

        final_id = manual_id or visitor_id
        return redirect("finish_registration_trusted", visitor_id=final_id)

    return render(request, "kiosk/signature_trusted.html", {"visitor_id": visitor_id})


def finish_registration_trusted_view(request, visitor_id):
    manual_id = request.session.get("manual_visitor_id")
    final_id = manual_id or visitor_id

    visitor = (Visitor.objects
               .filter(visitor_id=final_id)
               .order_by('-start_time')
               .first())
    if not visitor:
        return render(request, 'kiosk/error.html', {'message': _("Visitor not found.")})

    if not visitor.language:
        visitor.language = request.session.get(settings.LANGUAGE_COOKIE_NAME)

    # ✅ zatwierdź od razu (po chipie nie trafia do Pending Approvals)
    visitor.approved = True
    visitor.known_guest = True

    # --- decydujemy, czy DRUKOWAĆ ---
    has_badge = request.session.get('has_badge', True)  # kiosku: zapisywane w received_badge_view
    # Drukujemy, gdy:
    #  - produkcja (zawsze etykieta), ALBO
    #  - nie-produkcyjny i NIE ma identyfikatora (naklejka jako ID)
    should_print = bool(visitor.production_area or (not visitor.production_area and not has_badge))

    # --- przygotuj ZPL (tylko jeśli trzeba) ---
    zpl_payload = None
    if should_print:
        try:
            template_path = os.path.join(settings.BASE_DIR, 'GuestBook', 'Print_templates', 'template_zebra.zpl')
            with open(template_path, encoding='utf-8') as tpl:
                zpl_template = tpl.read()
            zpl_payload = zpl_template.format(
                company=_company_display(visitor),
                first_name=visitor.first_name,
                last_name=visitor.last_name,
                visit_purpose=visitor.visit_purpose or '',
                supervisor=visitor.host.host_name if visitor.host else ''
            )
            visitor.print_status = "pending"
        except Exception:
            # nie udało się przygotować szablonu – oznacz błąd druku
            visitor.print_status = "error"
    else:
        visitor.print_status = "skipped"

    # --- Email do gospodarza (nie blokuj widoku) ---
    host_email = visitor.host.email if visitor.host else ''
    status_alergen = "Uwaga - Gość jest uczulony na jeden lub więcej naszych alergenów." \
        if _is_yes(visitor.safety_question_3) else ""
    message = (
        f"Informacja o przybyciu gościa.\n\n"
        f"Dane:\n"
        f"Imię i nazwisko: {visitor.first_name} {visitor.last_name}\n"
        f"Telefon: {visitor.phone}\n"
        f"Firma: {_company_display(visitor)}\n"
        f"Cel wizyty: {visitor.visit_purpose}\n\n"
        f"Gość przebywa na obszarze zakładu.\n\n"
        f"{status_alergen}"
    )
    visitor.sms_status = "pending"

    # ✅ zapisz bieżący stan zanim odpalimy wątki
    visitor.save(update_fields=[
        'approved', 'known_guest', 'language', 'print_status', 'sms_status'
    ])

    # --- odpal wątki (fire & forget) ---
    if should_print and zpl_payload and visitor.print_status == "pending":
        from .models import KioskSettings
        _ks = KioskSettings.get()
        print_badge_async(visitor.id, zpl_payload, soft_timeout=5,
                          printer_ip=_ks.printer_address, printer_port=_ks.printer_port)

    send_email_to_host(visitor.id, host_email, 'New visitor arrived', message)

    # --- kolor etykiety na ekranie (informacyjnie) ---
    show_badge = should_print
    if show_badge:
        # Czerwony TYLKO: produkcja + z nadzorem
        # Zielony: produkcja + bez nadzoru  LUB  nie-produkcja bez ID (druk)
        if visitor.production_area and visitor.with_supervision:
            badge_color = _("Red (moving around the factory only with a host)")
            badge_css = "red"
        else:
            badge_color = _("Green (moving around the factory without a host)")
            badge_css = "green"
    else:
        badge_color = None
        badge_css = None

    # 🧹 sprzątanie sesji
    for key in [
        'trusted_visitor', 'visitor_data', 'badge_id', 'registration_mode',
        'safety_answers', 'safety_acknowledged', 'manual_visitor_id', 'has_badge'
    ]:
        request.session.pop(key, None)

    # ⛳️ natychmiast zwracamy widok – bez czekania na bramkę/drukarkę
    return render(request, 'kiosk/complete_trusted.html', {
        'badge_color': badge_color,
        'badge_css': badge_css,
        'is_production': visitor.production_area,
        'show_badge': show_badge,
    })

import threading
import logging
logger = logging.getLogger(__name__)

# def send_sms_with_timeout(number: str, message: str, timeout: int = 5) -> str:  # SMS disabled
#     ...replaced by send_email_with_timeout...

def send_email_with_timeout(email: str, subject: str, message: str, timeout: int = 5) -> str:
    """Send email with timeout. Returns 'sent' | 'error' | 'timeout' | 'skipped'."""
    from django.core.mail import send_mail
    from django.conf import settings as _s

    if not email:
        return 'skipped'

    result = {'status': 'timeout'}

    def _worker():
        try:
            send_mail(subject, message, _s.DEFAULT_FROM_EMAIL, [email], fail_silently=False)
            result['status'] = 'sent'
        except Exception as e:
            logger.exception('[EMAIL ERROR] %s', e)
            result['status'] = 'error'

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout)
    if t.is_alive():
        logger.warning('[EMAIL TIMEOUT] No response within %ss (to: %s)', timeout, email)
        return 'timeout'
    return result['status']


# utils_timeout.py (lub inny plik helpers)
from threading import Thread
from django.contrib.auth.models import Group
from .models import AdminProfile  # dostosuj ścieżkę importu do swojej struktury


def send_group_email_async(group_name: str, subject: str, text: str, timeout: int = 5):
    """Send email to all users in a Django group (replaces group SMS)."""
    from django.core.mail import send_mail
    from django.conf import settings as _s

    try:
        group = Group.objects.get(name=group_name)
    except Group.DoesNotExist:
        return

    emails = list(
        AdminProfile.objects
        .filter(user__in=group.user_set.all())
        .exclude(email='')
        .exclude(email='-')
        .values_list('email', flat=True)
    )

    for addr in set(emails):
        def _send(a=addr):
            try:
                send_mail(subject, text, _s.DEFAULT_FROM_EMAIL, [a], fail_silently=True)
            except Exception:
                pass
        Thread(target=_send, daemon=True).start()



from django.utils import timezone
from django.db.models import Q

def exit_badge_view(request):
    if request.method == 'POST':
        badge_id = request.POST.get('badge_id', '').strip()
        now = timezone.localtime()
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day   = now.replace(hour=23, minute=59, second=59, microsecond=999999)

        visitor = (Visitor.objects
                   .filter(visitor_id=badge_id, end_time__isnull=True,
                           start_time__range=(start_of_day, end_of_day))
                   .order_by('-start_time').first())

        if not visitor:
            return render(request, 'kiosk/exit_badge.html', {
                'error': _("No active visit found for this badge today. If this is a mistake, please contact reception.")
            })

        visitor.end_time = timezone.now()
        visitor.badge_returned = True
        visitor.save()

        return redirect('exit_done', visitor_id=badge_id)

    return render(request, 'kiosk/exit_badge.html')


def exit_done_view(request, visitor_id=None):
    visitor = (Visitor.objects
               .filter(visitor_id=visitor_id)
               .order_by('-start_time').first())

    if not visitor:
        return render(request, 'kiosk/exit_done.html', {'info': _("Visit closed.")})

    # Email do recepcji (grupa GuestBook_Reception), fire-and-forget
    try:
        msg = (
            f"Informacja: gość {visitor.first_name} {visitor.last_name} zakończył wizytę.\n"
            f"ID: {visitor.visitor_id}"
        )
        send_group_email_async('GuestBook_Reception', 'Visitor exit notification',
                               f"Guest {visitor.first_name} {visitor.last_name} has ended their visit. Badge ID: {visitor.visitor_id}")
    except Exception as e:
        print(f"[EMAIL EXIT FATAL] {e}")

    return render(request, 'kiosk/exit_done.html')


# dla kodu

def enter_code_view(request):
    if request.method == 'POST':
        code_input = request.POST.get('code')

        try:
            reservation_code = ReservationCode.objects.get(code=code_input)
        except ReservationCode.DoesNotExist:
            return render(request, 'kiosk/enter_code.html', {
                'error': _("Invalid code. Please try again.")
            })

        # sprawdzamy limit użyć
        if not reservation_code.can_use():
            return render(request, 'kiosk/enter_code.html', {
                'error': _("This code has already been used the maximum number of times.")
            })

        # ✅ zarejestruj użycie
        reservation_code.register_use()

        # powiązana rezerwacja
        reservation = reservation_code.reservation

        # zapisz w sesji
        request.session['reservation_id'] = reservation.id
        request.session['registration_mode'] = 'code'

        # 🔸 zawsze kieruj do wpisania identyfikatora
        return redirect('enter_visitor_id_for_code', visitor_id=reservation.id)

    return render(request, 'kiosk/enter_code.html')



def enter_visitor_id_for_code(request, visitor_id):
    if request.method == 'POST':
        manual_id = request.POST.get('visitor_id')
        request.session['manual_visitor_id'] = manual_id
        request.session['visitor_id'] = manual_id

        # 🔁 Jeśli wymagane jest BHP, przejdź najpierw do safety
        reservation = Reservation.objects.filter(id=visitor_id).first()
        return redirect('safety_from_code') if reservation and reservation.factory else redirect('signature_from_code')

    return render(request, 'kiosk/enter_id_for_code.html')


def signature_from_code_view(request):
    reservation_id = request.session.get('reservation_id')
    if not reservation_id:
        return redirect('enter_code')

    reservation = Reservation.objects.filter(id=reservation_id).first()
    if not reservation:
        return redirect('enter_code')

    if request.method == 'POST':
        data_url = request.POST.get('signature_data')
        format, imgstr = data_url.split(';base64,')
        data = ContentFile(base64.b64decode(imgstr), name=f"RES-{reservation_id}_signature.png")
        path = default_storage.save(f'signatures/{data.name}', data)
        request.session['signature_path'] = path

        return redirect('finalize_from_code')  # 🔁 ZAWSZE do finalize

    return render(request, 'kiosk/signature_from_code.html', {
        'reservation': reservation
    })


def safety_from_code_view(request):
    reservation_id = request.session.get('reservation_id')
    if not reservation_id:
        return redirect('enter_code')

    if request.method == 'POST':
        return redirect('bhp_question1_from_code')

    return render(request, 'kiosk/safety_image_only.html', {
        'visitor_id': reservation_id
    })


def bhp_question1_from_code_view(request):
    if request.method == 'POST':
        q1 = request.POST.get("question_1")

        if q1.lower() in ["tak", "yes"]:
            request.session.flush()
            return render(request, "kiosk/access_denied.html")

        request.session["safety_answers"] = {"question_1": q1}
        return redirect("bhp_question2_from_code")

    return render(request, "kiosk/bhp_question1_from_code.html")

def bhp_question2_from_code_view(request):
    if "safety_answers" not in request.session:
        return redirect("bhp_question1_from_code")

    if request.method == 'POST':
        q2 = request.POST.get("question_2")
        safety_answers = request.session["safety_answers"]
        safety_answers["question_2"] = q2
        request.session["safety_answers"] = safety_answers  # 🔁 Nadpisz całość
        return redirect("bhp_question3_from_code")

    return render(request, "kiosk/bhp_question2_from_code.html")

def bhp_question3_from_code_view(request):
    if "safety_answers" not in request.session:
        return redirect("bhp_question1_from_code")

    if request.method == 'POST':
        q3 = request.POST.get("question_3")
        safety_answers = request.session["safety_answers"]
        safety_answers["question_3"] = q3
        request.session["safety_answers"] = safety_answers  # 🔁 Nadpisz całość
        return redirect("signature_from_code")

    return render(request, "kiosk/bhp_question3_from_code.html")


from django.db import transaction
from django.db.models import F
from django.conf import settings
from django.utils.translation import gettext as _
import os

def _safe_visitor_id(reservation_pk: int) -> str:
    """3-znakowy identyfikator – np. modulo albo wycinek."""
    return str(reservation_pk)[-3:].zfill(3)

def _has_meaningful_changes(visitor, reservation, signature_path, safety_answers):
    """Zwraca True jeśli nowa finalizacja zmienia cokolwiek istotnego."""
    if visitor.production_area != reservation.factory:
        return True
    if visitor.with_supervision != reservation.supervision:
        return True
    if (visitor.safety_question_1 or "") != (safety_answers.get("question_1") or ""):
        return True
    if (visitor.safety_question_2 or "") != (safety_answers.get("question_2") or ""):
        return True
    if (visitor.safety_question_3 or "") != (safety_answers.get("question_3") or ""):
        return True
    # jeżeli zmienił się podpis (inny plik/ścieżka) – też licz jako zmianę
    if signature_path and str(visitor.signed) != str(signature_path):
        return True
    return False


@login_required
def finalize_from_code_view(request):
    reservation_id = request.session.get('reservation_id')
    signature_path = request.session.get('signature_path')
    safety_answers = request.session.get('safety_answers', {})

    if not reservation_id or not signature_path:
        return redirect('enter_code')

    reservation = Reservation.objects.filter(id=reservation_id).select_related('company', 'host').first()
    if not reservation:
        return redirect('enter_code')

    # przygotuj dane z sesji
    lang = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')
    visitor_id = request.session.get('visitor_id') or _safe_visitor_id(reservation.id)

    # Transakcja: nadpisanie Visitora + inkrement kodu
    try:
        with transaction.atomic():
            rc = ReservationCode.objects.select_for_update().get(reservation=reservation)

            # Pobierz JEDYNEGO Visitora dla tej rezerwacji i ZABLOKUJ
            visitor = (Visitor.objects
                       .select_for_update()
                       .filter(reservation=reservation)
                       .first())
            is_new = visitor is None

            # --- ZAPAMIĘTAJ POLA, KTÓRYCH NIE CHCESZ GUBIĆ ---
            keep_approved_by = visitor.approved_by if visitor else None
            keep_approved = visitor.approved if visitor else False
            keep_returned_by = visitor.returned_by if visitor else None
            keep_badge_returned = visitor.badge_returned if visitor else False

            if is_new:
                visitor = Visitor(
                    reservation=reservation,
                    visitor_id=(request.session.get('visitor_id') or _safe_visitor_id(reservation.id)),
                    first_name=reservation.visitor_first_name,
                    last_name=reservation.visitor_last_name,
                    phone=reservation.phone,
                    factory=reservation.company,
                    company_name_text=(reservation.company.company_name if reservation.company else ""),
                    visit_purpose=(reservation.other_purpose if reservation.visit_purpose_choice == 'inne'
                                   else reservation.visit_purpose_choice or ""),
                    host=reservation.host,
                    language=lang,
                    sms_status='pending',
                    print_status='pending',
                )

            # --- TU NADPISUJESZ DANE Z BIEŻĄCEJ REJESTRACJI ---
            visitor.production_area = reservation.factory
            visitor.with_supervision = reservation.supervision
            visitor.safety_acknowledged = True
            visitor.safety_question_1 = safety_answers.get("question_1")
            visitor.safety_question_2 = safety_answers.get("question_2")
            visitor.safety_question_3 = safety_answers.get("question_3")
            if signature_path:
                visitor.signed = signature_path

            visitor.print_status = 'pending'
            visitor.sms_status = 'pending'
            visitor.id_issued = False
            visitor.end_time = None

            # --- PRZYWRÓĆ ZAPAMIĘTANE POLA OPERACYJNE ---
            visitor.approved_by = keep_approved_by
            visitor.approved = keep_approved
            visitor.returned_by = keep_returned_by
            visitor.badge_returned = keep_badge_returned

            visitor.save()

            # Zlicz użycie (jeśli tak ma być)
            rc.usage_count = F('usage_count') + 1
            rc.save(update_fields=['usage_count'])

    except ReservationCode.DoesNotExist:
        pass

        # Sprzątanie sesji
    for key in ['reservation_id', 'signature_path', 'safety_answers', 'visitor_id']:
        request.session.pop(key, None)

    return redirect('complete_code', reservation_id=reservation.id)




def complete_code_view(request, reservation_id=None):
    """
    Finalizacja rejestracji opartej o REZERWACJĘ:
    - Aktualizuje JEDNEGO Visitora powiązanego z tą rezerwacją (FK).
    - Nie tworzy nowych wpisów.
    - Druk i SMS z limitem 5s i zapisem statusów.
    """
    # 1) Ustal rezerwację
    if reservation_id is None:
        reservation_id = request.session.get('reservation_id')
    reservation = get_object_or_404(Reservation, id=reservation_id)

    # 2) Znajdź istniejącego Visitora powiązanego z tą rezerwacją
    visitor = (Visitor.objects
               .filter(reservation=reservation, end_time__isnull=True)
               .order_by('-start_time')
               .first())

    if not visitor:
        # Jeżeli tu trafiasz, to w ścieżce rejestracji „na kod" nie zostało
        # przypisane visitor.reservation = reservation. Uzupełnij to w miejscu
        # tworzenia/aktualizacji Visitora (np. w widoku podpisu).
        return render(request, 'kiosk/error.html', {
            'message': _("Visitor not found or already completed.")
        })

    # 3) Uzupełnij język (jeśli brak)
    if not visitor.language:
        visitor.language = request.session.get(settings.LANGUAGE_COOKIE_NAME, 'pl')

    # 4) DRUK (max 5s) – drukujemy tylko, jeśli to ma sens (np. produkcja)
    try:
        if visitor.production_area:
            template_path = os.path.join(settings.BASE_DIR, 'GuestBook', 'Print_templates', 'template_zebra.zpl')
            with open(template_path, encoding='utf-8') as tpl:
                zpl_template = tpl.read()

            zpl_filled = zpl_template.format(
                company=_company_display(visitor),
                first_name=visitor.first_name,
                last_name=visitor.last_name,
                visit_purpose=visitor.visit_purpose or '',
                supervisor=visitor.host.host_name if visitor.host else ''
            )

            from .models import KioskSettings
            _ks = KioskSettings.get()
            pr_status = run_with_timeout(send_zpl_to_printer, zpl_filled,
                                         printer_ip=_ks.printer_address, port=_ks.printer_port, seconds=5)
            visitor.print_status = {"ok": "printed", "timeout": "timeout", "error": "error"}.get(pr_status, "error")
        else:
            visitor.print_status = "skipped"
    except Exception:
        visitor.print_status = "error"

    # 5) Email (max 5s) – do gospodarza (zastąpienie SMS)
    try:
        if visitor.host and visitor.host.email:
            status_msg = (
                "Gość oczekuje na odebranie z recepcji."
                if visitor.with_supervision else
                "Gość przebywa na obszarze zakładu."
            )
            status_alergen = (
                "Uwaga - Gość jest uczulony na jeden lub więcej naszych alergenów."
                if _is_yes(visitor.safety_question_3) else ""
            )
            message = (
                f"Informacja o przybyciu gościa.\n\n"
                f"Dane:\n"
                f"Imię i nazwisko: {visitor.first_name} {visitor.last_name}\n"
                f"Telefon: {visitor.phone}\n"
                f"Firma: {_company_display(visitor)}\n"
                f"Cel wizyty: {visitor.visit_purpose}\n\n"
                f"{status_msg}\n\n"
                f"{status_alergen}"
            )
            status = send_email_with_timeout(visitor.host.email, 'New visitor arrived', message, timeout=5)
            visitor.sms_status = {"sent": "sent", "timeout": "timeout", "error": "error"}.get(status, "error")
        else:
            visitor.sms_status = "skipped"
    except Exception:
        visitor.sms_status = "error"

    visitor.save()

    # 6) Kolor etykiety tylko informacyjnie na ekranie
    if visitor.production_area:
        if visitor.with_supervision:
            badge_color = _("Red (moving around the factory only with a host)")
            badge_css = "red"
        else:
            badge_color = _("Green (moving around the factory without a host)")
            badge_css = "green"
    else:
        badge_color = None
        badge_css = None

    return render(request, 'kiosk/complete_code.html', {
        'badge_color': badge_color,
        'badge_css': badge_css,
        'is_production': visitor.production_area,
    })


@login_required
def dashboard(request):
    query = request.GET.get("q", "").strip()
    guests = Visitor.objects.filter(approved=True)

    if query:
        query_lower = query.lower()
        guests = guests.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(factory__icontains=query) |
            Q(visit_purpose__icontains=query) |
            Q(host_name__icontains=query) |
            Q(host_phone__icontains=query) |
            Q(visitor_id__icontains=query) |
            (
                Q(production_area=True) if query_lower in ['tak', 'yes', 'produk', 'obszar'] else Q()
            ) |
            (
                Q(production_area=False) if query_lower in ['nie', 'no'] else Q()
            ) |
            (
                Q(with_supervision=True) if query_lower in ['z nadzorem', 'nadzor', 'tak', 'yes'] else Q()
            ) |
            (
                Q(with_supervision=False) if query_lower in ['bez nadzoru', 'nie', 'no'] else Q()
            )
        )

    guests = guests.order_by('-start_time')
    return render(request, 'panel/dashboard.html', {'guests': guests, 'query': query})


def nearby_persons(request):
    # Pokaż WSZYSTKICH obecnych (niezwrócony identyfikator + brak end_time)
    guests = (Visitor.objects
              .filter(approved=True, badge_returned=False, end_time__isnull=True)
              .order_by('-start_time'))
    return render(request, 'panel/nearby.html', {'guests': guests})


@login_required
def mark_returned(request, pk):
    guest = get_object_or_404(Visitor, pk=pk)
    guest.badge_returned = True
    guest.end_time = timezone.now()
    guest.returned_by = request.user
    guest.save()

    # 🔹 Zmień status rezerwacji na 'completed'
    if guest.reservation:
        guest.reservation.status = 'completed'
        guest.reservation.save()

    return redirect('nearby_persons')


@login_required
def statistics_view(request):
    today = timezone.now().date()
    week_number = today.isocalendar()[1]

    # Podstawowe liczniki
    month_visits = Visitor.objects.filter(start_time__month=today.month).count()
    week_visits = Visitor.objects.filter(start_time__week=week_number).count()
    day_visits = Visitor.objects.filter(start_time__date=today).count()

    # Obecni goście i statusy
    current_visitors = Visitor.objects.filter(end_time__isnull=True).count()
    production_now = Visitor.objects.filter(end_time__isnull=True, production_area=True).count()
    open_visits = Visitor.objects.filter(end_time__isnull=True).count()
    finished_visits = Visitor.objects.filter(badge_returned=True, end_time__isnull=False).count()
    auto_closed = Visitor.objects.filter(end_time__isnull=False, badge_returned=False).count()

    # Najczęstsze dane (w tym miesiącu)
    top_purpose = (
        Visitor.objects.filter(start_time__month=today.month)
        .values('visit_purpose')
        .annotate(total=Count('id'))
        .order_by('-total')
        .first()
    )
    top_host = (
        Visitor.objects.filter(start_time__month=today.month)
        .values('host__host_name')
        .annotate(total=Count('id'))
        .order_by('-total')
        .first()
    )

    # Rezerwacje (nadchodzące)
    upcoming_reservations = Reservation.objects.filter(date__gte=today, status='sent').count()

    # 🔥 NOWE: Nadchodzące rezerwacje per dzień (Top 7 dni)
    # Nadchodzące rezerwacje per dzień (Top 7 dni) – wersja bez TruncDate dla SQLite
    upcoming_per_day = (
        Reservation.objects
        .filter(date__gte=today, status='sent')
        .values('date')  # używamy bezpośrednio pola date
        .annotate(count=Count('id'))
        .order_by('date')[:7]
    )
    reservation_labels = [str(r['date']) for r in upcoming_per_day]
    reservation_values = [r['count'] for r in upcoming_per_day]

    # Production vs Non-production (ogółem)
    production_count = Visitor.objects.filter(production_area=True).count()
    non_production_count = Visitor.objects.filter(production_area=False).count()

    # Najczęściej odwiedzani hostowie (Top 5)
    top_hosts = (
        Visitor.objects.values('host__host_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )
    host_labels = [h['host__host_name'] or "Unknown" for h in top_hosts]
    host_values = [h['count'] for h in top_hosts]

    return render(request, 'panel/statistics.html', {
        'month': month_visits,
        'week': week_visits,
        'day': day_visits,
        'current_visitors': current_visitors,
        'production_now': production_now,
        'open_visits': open_visits,
        'finished': finished_visits,
        'auto_closed': auto_closed,
        'top_purpose': top_purpose['visit_purpose'] if top_purpose else 'No data',
        'top_host': top_host['host__host_name'] if top_host else 'No data',
        'upcoming_reservations': upcoming_reservations,

        # Chart.js
        'status_labels': json.dumps(["Open", "Finished", "Auto-closed"]),
        'status_data': json.dumps([open_visits, finished_visits, auto_closed]),
        'production_labels': json.dumps(["Production", "Non-production"]),
        'production_data': json.dumps([production_count, non_production_count]),
        'host_labels': json.dumps(host_labels),
        'host_values': json.dumps(host_values),

        # NOWE
        'reservation_labels': json.dumps(reservation_labels),
        'reservation_values': json.dumps(reservation_values),
    })

from django.core.mail import EmailMessage, get_connection
from django.conf import settings
import socket
import logging

logger = logging.getLogger(__name__)

@login_required
def profile_view(request):
    profile, _ = AdminProfile.objects.get_or_create(user=request.user)
    password_form = PasswordChangeForm(request.user)
    has_res_access = request.user.groups.filter(name="Recevio_User").exists()

    if request.method == 'POST':
        if 'signature_submit' in request.POST:
            signature = request.FILES.get('signature')
            if signature:
                profile.signature = signature
                profile.save()
                messages.success(request, 'Signature updated successfully.')
            else:
                messages.error(request, 'No signature uploaded.')

        elif 'password_submit' in request.POST:
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Password changed successfully.')
            else:
                messages.error(request, 'Error changing password.')

        elif 'printer_submit' in request.POST:
            printer_address = request.POST.get('printer_address', '').strip()
            printer_port_str = request.POST.get('printer_port', '9100').strip()
            try:
                profile.printer_address = printer_address
                profile.printer_port = int(printer_port_str)
                profile.save()
                messages.success(request, 'Printer settings saved.')
            except ValueError:
                messages.error(request, 'Invalid port number.')

        elif 'request_res_access' in request.POST:
            if has_res_access:
                messages.info(request, "You already have access to reservations.")
                return redirect("profile")

            user = request.user
            reply_email = (profile.email or user.email or "").strip()

            # --- adresy / konfiguracja
            to_addr   = getattr(settings, "HELPDESK_EMAIL", "wilga.it@brueggen.com")
            from_addr = (getattr(settings, "DEFAULT_FROM_EMAIL", "") or "").strip()

            # szybkie sanity-checki konfiguracji
            missing = []
            if not from_addr:                      missing.append("DEFAULT_FROM_EMAIL")
            if not getattr(settings, "EMAIL_HOST", None): missing.append("EMAIL_HOST")
            if not getattr(settings, "EMAIL_PORT", None): missing.append("EMAIL_PORT")
            if missing:
                messages.error(
                    request,
                    "Email is not configured: missing " + ", ".join(missing) + "."
                )
                return redirect("profile")

            subject = "Recevio – request for reservation access"
            body = (
                "A user requests access to the 'Reservations' module.\n\n"
                f"User: {user.get_full_name() or '-'} (login: {user.username})\n"
                f"Reply-to: {reply_email or '-'}\n"
                f"Phone: {profile.phone_number or '-'}\n\n"
                "Please verify and, if appropriate, add to group: GuestBook_User."
            )

            try:
                # Jawnie otwórz połączenie SMTP (łatwiej diagnozować)
                with get_connection(fail_silently=False) as conn:
                    email = EmailMessage(
                        subject=subject,
                        body=body,
                        from_email=from_addr,              # NADAWCA z settings!
                        to=[to_addr],
                        reply_to=[reply_email] if reply_email else None,
                        headers={"X-Recevio-User": user.username},
                        connection=conn,
                    )
                    sent = email.send()  # 1 = OK, 0 = serwer nie przyjął

            except (socket.gaierror, ConnectionRefusedError) as e:
                logger.exception("SMTP connection error")
                messages.error(request, "Cannot connect to SMTP server. Check EMAIL_HOST/PORT.")
                return redirect("profile")
            except Exception as e:
                logger.exception("Sending helpdesk email failed")
                msg = "Could not send the request (SMTP error)."
                if getattr(settings, "DEBUG", False):
                    msg += f" {e.__class__.__name__}: {e}"
                messages.error(request, msg)
                return redirect("profile")

            if sent == 1:
                messages.success(request, "Request sent to Helpdesk. We'll contact you soon.")
            else:
                # brak wyjątku, ale serwer nie zaakceptował żadnego odbiorcy
                messages.error(
                    request,
                    "Email not accepted by SMTP server (send() returned 0). "
                    "Verify sender/recipient addresses and SMTP policy."
                )

            return redirect("profile")

    return render(request, "panel/profile.html", {
        "profile": profile,
        "has_res_access": has_res_access,
        "form": password_form,
    })

@login_required
def export_visitors_excel(request):
    start_date = parse_date(request.GET.get("start_date", ""))
    end_date = parse_date(request.GET.get("end_date", ""))
    includes = request.GET.getlist("include")
    statuses = request.GET.getlist("status")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # usuń domyślną pustą

    # Filtr VISITORS
    if "visitors" in includes:
        ws = wb.create_sheet("Visitors")
        headers = [
            'ID', 'First Name', 'Last Name', 'Phone', 'Company', 'Visit Purpose', 'Host',
            'Host Phone', 'Start Time', 'End Time', 'Production Area', 'With Supervision',
            'Language', 'BHP Q1', 'BHP Q2', 'BHP Q3', 'Badge Returned?'
        ]
        ws.append(headers)

        visitors = Visitor.objects.all()

        # Zakres dat
        if start_date:
            visitors = visitors.filter(start_time__date__gte=start_date)
        if end_date:
            visitors = visitors.filter(start_time__date__lte=end_date)

        # Statusy
        if statuses:
            q_filter = Q()
            if "open" in statuses:
                q_filter |= Q(end_time__isnull=True)
            if "finished" in statuses:
                q_filter |= Q(badge_returned=True, end_time__isnull=False)
            if "auto" in statuses:
                q_filter |= Q(badge_returned=False, end_time__isnull=False)
            visitors = visitors.filter(q_filter)

        for v in visitors:
            ws.append([
                v.visitor_id,
                v.first_name,
                v.last_name,
                v.phone,
                str(v.factory) if v.factory else "-",
                v.visit_purpose,
                v.host.host_name if v.host else "",
                v.host.phone if v.host else "",
                v.start_time.strftime('%d.%m.%Y %H:%M') if v.start_time else '',
                v.end_time.strftime('%d.%m.%Y %H:%M') if v.end_time else '',
                'Yes' if v.production_area else 'No',
                'Yes' if v.with_supervision else 'No',
                v.language,
                v.safety_question_1,
                v.safety_question_2,
                v.safety_question_3,
                'Yes' if v.badge_returned else 'No'
            ])

    # Filtr RESERVATIONS
    if "reservations" in includes:
        ws = wb.create_sheet("Reservations")
        ws.append([
            'Visitor Name', 'Company', 'Host', 'Phone', 'Date', 'Time',
            'Purpose', 'Status'
        ])
        reservations = Reservation.objects.all()
        if start_date:
            reservations = reservations.filter(date__gte=start_date)
        if end_date:
            reservations = reservations.filter(date__lte=end_date)

        for r in reservations:
            ws.append([
                f"{r.visitor_first_name} {r.visitor_last_name}",
                str(r.company) if r.company else "",
                r.host.host_name if r.host else "",
                r.phone,
                r.date.strftime("%d.%m.%Y"),
                r.time.strftime("%H:%M") if r.time else "",
                r.other_purpose or r.visit_purpose_choice,
                r.status
            ])

    # Filtr TRUSTED VISITORS
    if "trusted" in includes:
        ws = wb.create_sheet("Trusted Visitors")
        ws.append([
            'First Name', 'Last Name', 'Phone', 'Company', 'Visit Purpose',
            'Host', 'Host Phone', 'Language', 'Production Area', 'With Supervision'
        ])
        trusted = TrustedVisitor.objects.all()
        for t in trusted:
            ws.append([
                t.first_name,
                t.last_name,
                t.phone,
                t.company,
                t.visit_purpose,
                t.host_name,
                t.host_phone,
                t.language,
                'Yes' if t.production_area else 'No',
                'Yes' if t.with_supervision else 'No'
            ])

    # Ustaw szerokość kolumn
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Generuj plik
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    today = datetime.now().strftime('%Y-%m-%d_%H-%M')
    response['Content-Disposition'] = f'attachment; filename=visitors_{today}.xlsx'
    wb.save(response)
    return response



@login_required
def visitor_detail(request, pk):
    guest = get_object_or_404(Visitor, pk=pk)

    if request.method == "POST" and "update_end_time" in request.POST:
        new_end_time = request.POST.get("end_time")
        if new_end_time:
            try:
                # Zapisujemy nową datę i godzinę
                guest.end_time = timezone.datetime.strptime(new_end_time, "%Y-%m-%dT%H:%M")
                guest.save()
                messages.success(request, "End Time updated successfully.")
                return redirect("visitor_detail", pk=guest.pk)
            except ValueError:
                messages.error(request, "Invalid date format. Please use the picker.")
        else:
            guest.end_time = None
            guest.save()
            messages.info(request, "End Time cleared.")

    return render(request, "panel/visitor_detail.html", {"guest": guest})


@login_required
def visitor_edit_view(request, pk):
    visitor = get_object_or_404(Visitor, pk=pk)

    # Dostęp tylko dla użytkowników z grupy GuestBook_Helpdesk
    if not request.user.groups.filter(name="Recevio_Helpdesk").exists():
        messages.error(request, "You do not have permission to edit this visitor.")
        return redirect("visitor_detail", pk=pk)

    if request.method == "POST":
        # Dane osobowe
        visitor.first_name = request.POST.get("first_name", "").strip()
        visitor.last_name = request.POST.get("last_name", "").strip()
        visitor.phone = request.POST.get("phone", "").strip()

        # Firma i host
        factory_id = request.POST.get("factory")
        visitor.factory = Company.objects.filter(id=factory_id).first() if factory_id else None

        host_id = request.POST.get("host")
        visitor.host = Host.objects.filter(id=host_id).first() if host_id else None

        # Production Area / With Supervision
        visitor.production_area = bool(request.POST.get("production_area"))
        visitor.with_supervision = bool(request.POST.get("with_supervision"))

        # Safety Questions
        visitor.safety_question_1 = request.POST.get("safety_question_1", "").strip()
        visitor.safety_question_2 = request.POST.get("safety_question_2", "").strip()
        visitor.safety_question_3 = request.POST.get("safety_question_3", "").strip()

        # Daty audytu
        audit_start_date = request.POST.get("audit_start_date")
        audit_end_date = request.POST.get("audit_end_date")
        visitor.audit_start_date = audit_start_date if audit_start_date else None
        visitor.audit_end_date = audit_end_date if audit_end_date else None

        # End Time
        end_time = request.POST.get("end_time")
        visitor.end_time = end_time if end_time else None

        # Approved / Badge Returned
        visitor.approved = bool(request.POST.get("approved"))
        visitor.badge_returned = bool(request.POST.get("badge_returned"))

        visitor.save()
        messages.success(request, "Visitor details updated successfully.")
        return redirect("visitor_detail", pk=pk)

    return render(request, "panel/visitor_edit.html", {
        "visitor": visitor,
        "companies": Company.objects.all(),
        "hosts": Host.objects.all()
    })


def generate_bhp_pdf_view(request, pk):
    guest = get_object_or_404(Visitor, pk=pk)

    lang = guest.language if hasattr(guest, 'language') and guest.language in ['pl', 'en'] else 'pl'

    pdf = generate_bhp_pdf(guest)

    return FileResponse(
        pdf,
        as_attachment=True,
        filename=f"BHP_{guest.first_name}_{guest.last_name}_{guest.start_time:%Y-%m-%d_%H-%M}.pdf"
    )

# views.py
from datetime import timedelta
from django.db.models import Q

# views.py
from datetime import timedelta
from django.db.models import Q
from django.utils import timezone

@login_required
def coming_visitors_view(request):
    today = timezone.localdate()
    cutoff = timezone.now() - timedelta(hours=24)

    reservations = (
        Reservation.objects
        .select_related('company','host','reservationcode')
        .filter(
            # przyszłe i NIE anulowane
            (Q(date__gte=today) & ~Q(status='cancelled'))
            # ALBO anulowane w ciągu 24h (niezależnie od daty)
            | Q(status='cancelled', cancelled_at__gte=cutoff)
        )
        .exclude(visitor__isnull=False)
        .order_by('date','time')
    )

    return render(request, 'panel/coming_visitors.html', {
        'reservations': reservations,
        'cutoff': cutoff,   # << do szablonu – aby warunkowo odblokować Resend
    })



from django.utils import timezone

@login_required
def reservation(request):
    profile = AdminProfile.objects.filter(user=request.user).first()

    reservations = (
        Reservation.objects
        .filter(user=request.user)
        .select_related('company', 'host', 'reservationcode')
        .order_by('-date', '-time')
    )

    error_reservation = None
    if not profile or not profile.signature:
        error_reservation = _("To make a reservation, you must first add a signature to your profile.")

    return render(request, 'panel/reservation.html', {
        'reservations': reservations,
        'editable_statuses': ['sent', 'arrived', 'timeout', 'error', 'no_number', 'pending'],  # ⬅️ + arrived
        'cancellable_statuses': ['sent', 'pending', 'timeout', 'error', 'no_number'],
        'error_reservation': error_reservation,
    })




# views.py
@login_required
def reservation_create_view(request):
    profile = AdminProfile.objects.filter(user=request.user).first()
    if not profile or not profile.signature:
        return redirect('reservation')

    if request.method == 'POST':
        form = ReservationForm(request.POST, disable_translation=True)
        if form.is_valid():
            reservation = form.save(commit=False)

            # firma
            if form.cleaned_data.get('company'):
                reservation.company = form.cleaned_data['company']
            elif form.cleaned_data.get('other_company'):
                company_obj, _ = Company.objects.get_or_create(
                    company_name=form.cleaned_data['other_company'].strip()
                )
                reservation.company = company_obj
            elif form.cleaned_data.get('no_company'):
                reservation.company = None

            reservation.user = request.user
            reservation.status = 'sent'
            reservation.sms_status = 'pending'   # ⬅️ od razu pokaż w UI „Pending"
            reservation.save()

            code = ReservationCode.objects.create(reservation=reservation)

            # SMS (max 5s) – DO KOGO WYSYŁAMY?
            try:
                # Jeżeli ma iść do gospodarza:
                number = reservation.host.phone if (reservation.host and reservation.host.phone) else None
                # Jeśli zamiast tego do gościa, użyj:
                # number = reservation.phone

                if not number:
                    reservation.sms_status = 'no_number'
                else:
                    message = (
                        f"Brüggen Polska.\n\n"
                        f"Rezerwacja wizyty została wykonana przez {reservation.host}.\n\n"
                        f"Dane gościa:\n"
                        f"Imię i nazwisko: {reservation.visitor_first_name} {reservation.visitor_last_name}\n"
                        f"Telefon: {reservation.phone}\n"
                        f"Firma: {reservation.company or 'Brak'}\n"
                        f"Cel wizyty: {reservation.other_purpose if reservation.visit_purpose_choice == 'inne' else reservation.visit_purpose_choice}\n\n"
                        f"Wizyta została zarezerwowana na dzień {reservation.date} o godzinie {reservation.time or '--'}.\n\n"
                        f"Kod rezerwacji: {code.code}"
                    )

                    # ⬇️ twardy timeout 5s — UI nie wisi dłużej
                    status = send_sms_with_timeout(number, message, timeout=5)
                    # mapowanie na kolory w UI
                    reservation.sms_status = {
                        'sent': 'sent',
                        'timeout': 'timeout',
                        'error': 'error'
                    }.get(status, 'error')
            except Exception:
                reservation.sms_status = 'error'

            reservation.save(update_fields=['sms_status'])
            return redirect('reservation')
    else:
        form = ReservationForm(disable_translation=True)

    return render(request, 'panel/reservation_form.html', {
        'form': form,
        'hosts': Host.objects.all()
    })

from django.utils import timezone
from django.http import HttpResponseForbidden

@login_required
def reservation_edit_view(request, pk):
    reservation = get_object_or_404(Reservation, pk=pk)

    is_owner = reservation.user_id == request.user.id
    is_reception = request.user.groups.filter(name="Recevio_Reception").exists()
    if not (is_owner or is_reception):
        return HttpResponseForbidden("You don't have permission to edit this reservation.")

    # Edytowalne dopóki NIE jest 'completed'
    EDITABLE_STATUSES = {'sent', 'arrived', 'timeout', 'error', 'no_number', 'pending'}

    if reservation.status not in EDITABLE_STATUSES:
        messages.error(request, "You can only edit non-completed reservations.")
        return redirect('reservation')

    rc = getattr(reservation, 'reservationcode', None)
    if rc and rc.usage_count >= rc.max_uses:
        messages.warning(request, "The reservation code has no remaining uses.")

    if request.method == 'POST':
        form = ReservationForm(request.POST, instance=reservation, disable_translation=True)
        if form.is_valid():
            reservation = form.save(commit=False)

            # Firma
            if form.cleaned_data.get('company'):
                reservation.company = form.cleaned_data['company']
            elif form.cleaned_data.get('other_company'):
                name = form.cleaned_data['other_company'].strip()
                company_obj, _ = Company.objects.get_or_create(company_name=name)
                reservation.company = company_obj
            elif form.cleaned_data.get('no_company'):
                reservation.company = None

            # (Uwaga: w Twoich modelach Host NIE ma pola company.
            # Jeżeli chcesz auto-host z Company, użyj Company.host_name)
            if not form.cleaned_data.get('host') and reservation.company and reservation.company.host_name:
                reservation.host = reservation.company.host_name

            reservation.save()
            messages.success(request, "Reservation updated successfully.")
            return redirect('reservation')
    else:
        form = ReservationForm(instance=reservation, disable_translation=True)

    return render(request, 'panel/reservation_form.html', {
        'form': form,
        'edit_mode': True,
        'reservation_id': reservation.id
    })


from django.db import transaction
from django.utils.translation import gettext as _
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
from django.views.decorators.http import require_POST

CANCELLABLE_STATUSES = {'sent', 'pending', 'timeout', 'error', 'no_number'}

@require_POST
@login_required
def reservation_cancel_view(request, pk):
    reservation = get_object_or_404(Reservation, pk=pk)

    # uprawnienia: właściciel lub recepcja
    is_owner = reservation.user_id == request.user.id
    is_reception = request.user.groups.filter(name="Recevio_Reception").exists()
    if not (is_owner or is_reception):
        return HttpResponseForbidden("You don't have permission to cancel this reservation.")

    # czy wolno anulować tę rezerwację?
    if reservation.status not in CANCELLABLE_STATUSES:
        messages.error(request, "You can only cancel non-arrived, non-completed reservations.")
        return redirect('reservation')

    # SMS do gościa (z twardym timeoutem)
    msg = (
        "Brüggen Polska.\n\n"
        f"Twoje spotkanie zaplanowane na {reservation.date.strftime('%d.%m.%Y')} "
        f"o {reservation.time or '--'} zostało ANULOWANE.\n\n"
        "W razie potrzeby prosimy o kontakt z osobą zapraszającą."
    )
    sms_status = send_sms_with_timeout(reservation.phone, msg, timeout=5)  # 'sent'|'timeout'|'error'|'no_number'

    # ustaw anulowanie + znacznik czasu
    reservation.status = 'cancelled'
    reservation.cancelled_at = timezone.now()  # <<<<<<
    reservation.sms_status = sms_status  # od razu odśwież „Send"
    reservation.save(update_fields=['status', 'cancelled_at', 'sms_status'])

    # komunikaty UI
    if sms_status == 'sent':
        messages.success(request, "Reservation cancelled and SMS sent to the guest.")
    elif sms_status == 'no_number':
        messages.warning(request, "Reservation cancelled. No phone number to notify the guest.")
    elif sms_status == 'timeout':
        messages.warning(request, "Reservation cancelled. SMS sending timed out.")
    else:
        messages.warning(request, "Reservation cancelled. Could not send SMS.")

    return redirect('reservation')




def guestbook(request):
    return render(request, 'panel/guestbook.html')


def reprint_badge_view(request, pk):
    visitor = get_object_or_404(Visitor, pk=pk)

    try:
        template_path = os.path.join(settings.BASE_DIR, 'GuestBook', 'Print_templates', 'template_zebra.zpl')
        with open(template_path, encoding='utf-8') as tpl:
            zpl_template = tpl.read()

        zpl_filled = zpl_template.format(
            company=visitor.factory.company_name if visitor.factory else '',
            first_name=visitor.first_name,
            last_name=visitor.last_name,
            visit_purpose=visitor.visit_purpose or '',
            supervisor=visitor.host.host_name if visitor.host else ''
        )

        _profile = AdminProfile.objects.filter(user=request.user).first() if request.user.is_authenticated else None
        _printer_ip = _profile.printer_address if _profile and _profile.printer_address else '10.30.40.150'
        _printer_port = _profile.printer_port if _profile else 9100
        send_zpl_to_printer(zpl_filled, printer_ip=_printer_ip, port=_printer_port)

        messages.success(request, _("Etykieta została ponownie wydrukowana."))
    except Exception as e:
        messages.error(request, f"Błąd podczas wydruku: {e}")

    return redirect(request.META.get("HTTP_REFERER", "dashboard"))

@login_required
def pending_approvals(request):
    unapproved = Visitor.objects.filter(approved=False).order_by('-start_time')
    return render(request, 'panel/pending_approvals.html', {'unapproved_visitors': unapproved})


@login_required
def approve_visitor(request, pk):
    visitor = get_object_or_404(Visitor, pk=pk)

    if request.method == 'POST':
        known = request.POST.get('known_guest') == 'True'
        visitor.approved = True
        visitor.known_guest = known
        visitor.approved_by = request.user
        visitor.save()

        # 🔹 Zmień status powiązanej rezerwacji na "arrived"
        if visitor.reservation:
            visitor.reservation.status = 'arrived'
            visitor.reservation.save()

        return redirect('pending_approvals')

    return render(request, 'panel/approve_visitor.html', {'visitor': visitor})


@login_required
def companies_view(request):
    query = request.GET.get("q", "")
    companies = Company.objects.select_related("host_name").all()

    if query:
        companies = companies.filter(
            Q(company_name__icontains=query) |
            Q(host_name__host_name__icontains=query)
        )

    return render(request, "panel/companies.html", {
        "companies": companies,
        "query": query
    })


@login_required
def company_add(request):
    hosts = Host.objects.all()
    if request.method == "POST":
        name = request.POST.get("company_name")
        host_id = request.POST.get("host_name")
        host = Host.objects.filter(id=host_id).first() if host_id else None
        Company.objects.create(company_name=name, host_name=host)
        messages.success(request, "Company added successfully.")
        return redirect("companies")
    return render(request, "panel/company_form.html", {"hosts": hosts})


@login_required
def company_edit(request, pk):
    company = get_object_or_404(Company, pk=pk)
    hosts = Host.objects.all()
    if request.method == "POST":
        company.company_name = request.POST.get("company_name")
        host_id = request.POST.get("host_name")
        company.host_name = Host.objects.filter(id=host_id).first() if host_id else None
        company.save()
        messages.success(request, "Company updated successfully.")
        return redirect("companies")
    return render(request, "panel/company_form.html", {"company": company, "hosts": hosts})


@login_required
def company_delete(request, pk):
    company = get_object_or_404(Company, pk=pk)
    company.delete()
    messages.success(request, "Company deleted successfully.")
    return redirect("companies")


@login_required
def company_import(request):
    if request.method == "POST" and request.FILES.get("file"):
        xlsx_file = request.FILES["file"]
        wb = openpyxl.load_workbook(xlsx_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            company_name = str(row[0]).strip() if row[0] else None
            host_name = str(row[1]).strip() if len(row) > 1 and row[1] else None

            if company_name:
                host = Host.objects.filter(host_name=host_name).first() if host_name else None
                Company.objects.get_or_create(company_name=company_name, host_name=host)

        messages.success(request, "Companies imported successfully from Excel.")
        return redirect("companies")
    return render(request, "panel/company_import.html")



@login_required
def hosts_view(request):
    query = request.GET.get("q", "")
    hosts = Host.objects.all()
    if query:
        hosts = hosts.filter(
            Q(host_name__icontains=query) | Q(phone__icontains=query)
        )
    return render(request, "panel/hosts.html", {"hosts": hosts, "query": query})


@login_required
def host_add(request):
    if request.method == "POST":
        Host.objects.create(
            host_name=request.POST.get("host_name"),
            phone=request.POST.get("phone"),
            email=request.POST.get("email", ""),
        )
        messages.success(request, "Host added successfully.")
        return redirect("hosts")
    return render(request, "panel/host_form.html")


@login_required
def host_edit(request, pk):
    host = get_object_or_404(Host, pk=pk)
    if request.method == "POST":
        host.host_name = request.POST.get("host_name")
        host.phone = request.POST.get("phone")
        host.email = request.POST.get("email", "")
        host.save()
        messages.success(request, "Host updated successfully.")
        return redirect("hosts")
    return render(request, "panel/host_form.html", {"host": host})


@login_required
def host_delete(request, pk):
    host = get_object_or_404(Host, pk=pk)
    host.delete()
    messages.success(request, "Host deleted successfully.")
    return redirect("hosts")


@login_required
def host_import(request):
    if request.method == "POST" and request.FILES.get("file"):
        xlsx_file = request.FILES["file"]
        wb = openpyxl.load_workbook(xlsx_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            host_name = str(row[0]).strip() if row[0] else None
            phone = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            email = str(row[2]).strip() if len(row) > 2 and row[2] else ""

            if host_name:
                host, created = Host.objects.get_or_create(host_name=host_name, defaults={"phone": phone, "email": email})
                if not created:
                    updated = False
                    if phone and host.phone != phone:
                        host.phone = phone
                        updated = True
                    if email and host.email != email:
                        host.email = email
                        updated = True
                    if updated:
                        host.save()

        messages.success(request, "Hosts imported successfully from Excel.")
        return redirect("hosts")
    return render(request, "panel/host_import.html")


@login_required
def trusted_view(request):
    query = request.GET.get("q", "")
    badge_filter = request.GET.get("badge", "")
    sort = request.GET.get("sort", "")
    trusted = TrustedVisitor.objects.all()

    if query:
        trusted = trusted.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(company__icontains=query) |
            Q(host_name__icontains=query)
        )

    if badge_filter:
        trusted = trusted.filter(badge_id__icontains=badge_filter)

    # ✅ Sortowanie – długie kody (alfanumeryczne i >3 cyfry) na górze, krótkie cyfry na dole
    if sort == "badge":
        trusted = trusted.annotate(
            is_short_numeric=Case(
                When(badge_id__regex=r'^\d{1,3}$', then=Value(True)),  # tylko 1-3 cyfry
                default=Value(False),
                output_field=BooleanField()
            )
        ).order_by("is_short_numeric", "badge_id")

    return render(request, "panel/trusted.html", {
        "trusted": trusted,
        "query": query,
        "badge_filter": badge_filter,
        "sort": sort
    })


@login_required
def trusted_add(request):
    if request.method == "POST":
        TrustedVisitor.objects.create(
            first_name=request.POST.get("first_name"),
            last_name=request.POST.get("last_name"),
            phone=request.POST.get("phone"),
            company=request.POST.get("company"),
            visit_purpose=request.POST.get("visit_purpose"),
            host_name=request.POST.get("host_name"),
            host_phone=request.POST.get("host_phone"),
            language=request.POST.get("language"),
            production_area=bool(request.POST.get("production_area")),
            with_supervision=bool(request.POST.get("with_supervision")),
            badge_id=request.POST.get("badge_id")
        )
        # messages.success(request, "Trusted visitor added successfully.")
        return redirect("trusted")
    return render(request, "panel/trusted_form.html")


@login_required
def trusted_edit(request, pk):
    trusted = get_object_or_404(TrustedVisitor, pk=pk)
    if request.method == "POST":
        trusted.first_name = request.POST.get("first_name")
        trusted.last_name = request.POST.get("last_name")
        trusted.phone = request.POST.get("phone")
        trusted.company = request.POST.get("company")
        trusted.visit_purpose = request.POST.get("visit_purpose")
        trusted.host_name = request.POST.get("host_name")
        trusted.host_phone = request.POST.get("host_phone")
        trusted.language = request.POST.get("language")
        trusted.production_area = bool(request.POST.get("production_area"))
        trusted.with_supervision = bool(request.POST.get("with_supervision"))
        trusted.badge_id = request.POST.get("badge_id")
        trusted.save()
        # messages.success(request, "Trusted visitor updated successfully.")
        return redirect("trusted")

    return render(request, "panel/trusted_form.html", {"trusted": trusted})


@login_required
def trusted_delete(request, pk):
    trusted = get_object_or_404(TrustedVisitor, pk=pk)
    trusted.delete()
    # messages.success(request, "Trusted visitor deleted successfully.")
    return redirect("trusted")


@login_required
def reservation_visitor_view(request):
    # Dostęp dla Reception i Helpdesk
    if not (
        request.user.groups.filter(name="Recevio_Reception").exists() or
        request.user.groups.filter(name="Recevio_Helpdesk").exists()
    ):
        messages.error(request, "You do not have permission to view reservations.")
        return redirect("dashboard")

    query = request.GET.get("q", "")
    status_filter = request.GET.get("status", "")

    reservations = Reservation.objects.all().select_related("company", "host", "user")

    if query:
        reservations = reservations.filter(
            Q(visitor_first_name__icontains=query) |
            Q(visitor_last_name__icontains=query) |
            Q(company__company_name__icontains=query) |
            Q(host__host_name__icontains=query) |
            Q(phone__icontains=query)
        )

    if status_filter:
        reservations = reservations.filter(status=status_filter)

    return render(request, "panel/reservation_visitor.html", {
        "reservations": reservations,
        "query": query,
        "status_filter": status_filter
    })

@login_required
def reservation_visitor_edit_view(request, pk):
    reservation = get_object_or_404(Reservation, pk=pk)

    # Dostęp tylko dla Reception i Helpdesk
    if not (
        request.user.groups.filter(name="Recevio_Reception").exists() or
        request.user.groups.filter(name="Recevio_Helpdesk").exists()
    ):
        messages.error(request, "You do not have permission to edit this reservation.")
        return redirect("reservation_visitor")

    if request.method == "POST":
        reservation.visitor_first_name = request.POST.get("visitor_first_name", "").strip()
        reservation.visitor_last_name = request.POST.get("visitor_last_name", "").strip()
        reservation.phone = request.POST.get("phone", "").strip()

        company_id = request.POST.get("company")
        reservation.company = Company.objects.filter(id=company_id).first() if company_id else None

        host_id = request.POST.get("host")
        reservation.host = Host.objects.filter(id=host_id).first() if host_id else None

        reservation.date = request.POST.get("date")
        reservation.status = request.POST.get("status")

        # Audit dates
        reservation.audit_start_date = request.POST.get("audit_start_date") or None
        reservation.audit_end_date = request.POST.get("audit_end_date") or None

        # Production area (factory) and supervision
        reservation.factory = bool(request.POST.get("factory"))
        reservation.supervision = bool(request.POST.get("supervision"))

        # Conference needed
        reservation.conference_needed = bool(request.POST.get("conference_needed"))
        reservation.conference_room = (
            request.POST.get("conference_room") if reservation.conference_needed else None
        )

        # Visit purpose (standard or other)
        reservation.visit_purpose_choice = request.POST.get("visit_purpose_choice", "")
        reservation.other_purpose = (
            request.POST.get("other_purpose", "").strip()
            if reservation.visit_purpose_choice == "other"
            else ""
        )

        reservation.save()
        messages.success(request, "Reservation updated successfully.")
        return redirect("reservation_visitor")

    return render(request, "panel/reservation_visitor_edit.html", {
        "reservation": reservation,
        "companies": Company.objects.all(),
        "hosts": Host.objects.all(),
    })


@login_required
def reservation_add_view(request):
    # Dostęp tylko dla Reception i Helpdesk
    if not (
        request.user.groups.filter(name="Recevio_Reception").exists() or
        request.user.groups.filter(name="Recevio_Helpdesk").exists()
    ):
        messages.error(request, "You do not have permission to add reservations.")
        return redirect("reservation_visitor")

    if request.method == "POST":
        visitor_first_name = request.POST.get("visitor_first_name", "").strip()
        visitor_last_name = request.POST.get("visitor_last_name", "").strip()
        phone = request.POST.get("phone", "").strip()

        company_id = request.POST.get("company")
        company = Company.objects.filter(id=company_id).first() if company_id else None

        host_id = request.POST.get("host")
        host = Host.objects.filter(id=host_id).first() if host_id else None

        date = request.POST.get("date")
        status = request.POST.get("status", "sent")

        audit_start_date = request.POST.get("audit_start_date") or None
        audit_end_date = request.POST.get("audit_end_date") or None

        factory = bool(request.POST.get("factory"))
        supervision = bool(request.POST.get("supervision"))

        conference_needed = bool(request.POST.get("conference_needed"))
        conference_room = request.POST.get("conference_room") if conference_needed else None

        visit_purpose_choice = request.POST.get("visit_purpose_choice", "")
        other_purpose = (
            request.POST.get("other_purpose", "").strip()
            if visit_purpose_choice == "other"
            else ""
        )

        # ✅ Tworzenie rezerwacji
        reservation = Reservation.objects.create(
            user=request.user,
            visitor_first_name=visitor_first_name,
            visitor_last_name=visitor_last_name,
            phone=phone,
            company=company,
            host=host,
            date=date,
            status=status,
            audit_start_date=audit_start_date,
            audit_end_date=audit_end_date,
            factory=factory,
            supervision=supervision,
            conference_needed=conference_needed,
            conference_room=conference_room,
            visit_purpose_choice=visit_purpose_choice,
            other_purpose=other_purpose,
        )

        # ✅ Automatyczne utworzenie kodu rezerwacji
        code = ReservationCode.objects.create(reservation=reservation)

        # ✅ Wysyłka SMS do gospodarza
        if reservation.host and reservation.host.phone:
            try:
                sms = SMSGateway()
                message = (
                    f"Brüggen Polska.\n\n"
                    f"Rezerwacja wizyty została dodana przez {request.user}.\n\n"
                    f"Dane gościa:\n"
                    f"Imię i nazwisko: {reservation.visitor_first_name} {reservation.visitor_last_name}\n"
                    f"Telefon: {reservation.phone}\n"
                    f"Firma: {reservation.company or 'Brak'}\n"
                    f"Cel wizyty: {reservation.other_purpose if reservation.visit_purpose_choice == 'other' else reservation.visit_purpose_choice}\n\n"
                    f"Wizyta została zarezerwowana na dzień {reservation.date}.\n"
                    f"Kod rezerwacji: {code.code}"
                )
                result = sms.send_sms(reservation.host.phone, message)
                if result.get("status") == "success":
                    print(f"[SMS SENT ✅] to {reservation.host.phone} | Response: {result}")
                else:
                    print(f"[SMS ERROR ❌] {result}")
            except Exception as e:
                print(f"[SMS FATAL ERROR] {e}")
        else:
            print("[SMS INFO] Brak numeru telefonu gospodarza.")

        messages.success(request, f"Reservation added successfully (Code: {code.code}).")
        return redirect("reservation_visitor")

    return render(request, "panel/reservation_add.html", {
        "companies": Company.objects.all(),
        "hosts": Host.objects.all(),
        "today": timezone.now().date()
    })

from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils.http import url_has_allowed_host_and_scheme
from django.db import transaction

# Jeżeli masz już to gdzieś zdefiniowane, użyj istniejącego:
# send_sms_with_timeout(number: str, text: str, timeout: int) -> 'sent' | 'timeout' | 'error' | 'no_number'

def _safe_next(request):
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or ''
    return next_url if url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}) else None

# views.py
from datetime import timedelta
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponseForbidden
from django.contrib import messages
from django.urls import reverse

# Zakładam, że masz tę funkcję:
# send_sms_with_timeout(number, text, timeout=5) -> 'sent' | 'timeout' | 'error' | 'no_number'

@require_POST
@login_required
def reservation_resend_sms_view(request, pk):
    # Pobranie rezerwacji
    reservation = get_object_or_404(
        Reservation.objects.select_related('host', 'reservationcode'),
        pk=pk
    )

    # Uprawnienia: właściciel LUB recepcja/helpdesk
    is_owner = reservation.user_id == request.user.id
    in_reception_or_helpdesk = request.user.groups.filter(
        name__in=["Recevio_Reception", "Recevio_Helpdesk"]
    ).exists()
    if not (is_owner or in_reception_or_helpdesk):
        return HttpResponseForbidden("You don't have permission to resend SMS for this reservation.")

    # Dokąd wrócić po akcji
    next_url = request.POST.get("next") or reverse("coming_visitors")

    # Nie pozwalaj na resend dla COMPLETED
    if reservation.status == "completed":
        messages.error(request, "Cannot resend SMS for a completed reservation.")
        return redirect(next_url)

    # Ustal numer i treść, zależnie od statusu
    # 1) Dla anulowanych — wyślij ponownie SMS o ANULOWANIU do GOŚCIA, ale tylko w 24h od cancelled_at
    if reservation.status == "cancelled":
        # kontrola 24h
        if not reservation.cancelled_at or timezone.now() - reservation.cancelled_at > timedelta(hours=24):
            messages.error(request, "Cancelled reservation can be notified only within 24 hours of cancellation.")
            return redirect(next_url)

        number = reservation.phone or None  # do gościa
        msg = (
            "Brüggen Polska.\n\n"
            f"Twoje spotkanie zaplanowane na {reservation.date.strftime('%d.%m.%Y')} "
            f"o {reservation.time or '--'} zostało ANULOWANE.\n\n"
            "W razie potrzeby prosimy o kontakt z osobą zapraszającą."
        )

    # 2) Dla pozostałych — wyślij ponownie SMS z danymi rezerwacji do GOSPODARZA
    else:
        number = reservation.host.phone if (reservation.host and reservation.host.phone) else None
        code = getattr(reservation, "reservationcode", None)
        code_str = code.code if code else "—"
        purpose = reservation.other_purpose if reservation.visit_purpose_choice in ("inne", "other") \
                  else (reservation.visit_purpose_choice or "")

        msg = (
            "Brüggen Polska.\n\n"
            f"Rezerwacja wizyty – przypomnienie.\n\n"
            f"Dane gościa:\n"
            f"Imię i nazwisko: {reservation.visitor_first_name} {reservation.visitor_last_name}\n"
            f"Telefon: {reservation.phone}\n"
            f"Firma: {reservation.company if reservation.company else 'Brak'}\n"
            f"Cel wizyty: {purpose}\n\n"
            f"Termin: {reservation.date} {reservation.time or '--'}\n"
            f"Kod rezerwacji: {code_str}"
        )

    # Wyślij z twardym timeoutem 5s
    result = send_sms_with_timeout(number, msg, timeout=5)  # 'sent' | 'timeout' | 'error' | 'no_number'

    # Zaktualizuj status wysyłki (pokazujesz go w kolumnie "Send")
    reservation.sms_status = {
        'sent': 'sent',
        'timeout': 'timeout',
        'error': 'error',
        'no_number': 'no_number',
    }.get(result, 'error')
    reservation.save(update_fields=['sms_status'])

    # Komunikaty dla UI
    if result == 'sent':
        messages.success(request, "SMS has been resent successfully.")
    elif result == 'no_number':
        messages.warning(request, "Cannot resend SMS: no phone number available.")
    elif result == 'timeout':
        messages.warning(request, "Resend attempted, but the SMS gateway timed out.")
    else:
        messages.error(request, "SMS resend failed due to a gateway error.")

    return redirect(next_url)



# views.py
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.db.models import Q
from django.shortcuts import get_object_or_404, render, redirect

ALLOWED_GROUPS = (
    "Recevio_User",
    "Recevio_Reception",
    "Recevio_Helpdesk",
    "Recevio_BoxFlow",
)

def _require_helpdesk(user) -> bool:
    return user.groups.filter(name="Recevio_Helpdesk").exists() or user.is_superuser


@login_required
def helpdesk_user_list(request):
    if not _require_helpdesk(request.user):
        messages.error(request, "You do not have permission to view this page.")
        return redirect("dashboard")

    q = request.GET.get("q", "").strip()

    users = (User.objects
             .select_related()
             .prefetch_related("groups")
             .order_by("username"))

    if q:
        users = users.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q)
        )

    # nie pokazuj superuserów, żeby przypadkiem nie grzebać
    users = users.exclude(is_superuser=True)[:200]

    return render(request, "panel/helpdesk_user_list.html", {
        "users": users,
        "query": q,
        "allowed_groups": ALLOWED_GROUPS,
    })


@login_required
def helpdesk_user_edit(request, uid: int):
    if not _require_helpdesk(request.user):
        messages.error(request, "You do not have permission to edit users.")
        return redirect("dashboard")

    user = get_object_or_404(User.objects.prefetch_related("groups"), pk=uid)

    if user.is_superuser:
        messages.warning(request, "Superuser cannot be edited here.")
        return redirect("helpdesk_users")

    allowed_qs = Group.objects.filter(name__in=ALLOWED_GROUPS).order_by("name")
    current_allowed = set(user.groups.filter(name__in=ALLOWED_GROUPS))

    if request.method == "POST":
        # Zaznaczone checkboxy
        selected_ids = request.POST.getlist("groups")
        wanted = set(allowed_qs.filter(id__in=selected_ids))

        # Usuń tylko te dozwolone, których nie ma na liście
        for g in (current_allowed - wanted):
            user.groups.remove(g)
        # Dodaj brakujące
        for g in (wanted - current_allowed):
            user.groups.add(g)

        # (opcjonalnie) nie pozwól samemu sobie odebrać Helpdesku
        if user.pk == request.user.pk and not user.groups.filter(name="Recevio_Helpdesk").exists():
            # przywróć
            helpdesk = Group.objects.filter(name="Recevio_Helpdesk").first()
            if helpdesk:
                user.groups.add(helpdesk)
            messages.warning(request, "You cannot remove your own Helpdesk access. It has been restored.")

        messages.success(request, "User groups updated.")
        return redirect("helpdesk_users")

    return render(request, "panel/helpdesk_user_edit.html", {
        "edited": user,
        "allowed_groups": allowed_qs,
        "current_allowed_ids": {g.id for g in current_allowed},
    })


# BoxFlow logic

def _can_boxflow(user):
    return user.is_authenticated and user.groups.filter(name__in=["Recevio_BoxFlow"]).exists()

def _can_helpdesk(user):
    return user.is_authenticated and user.groups.filter(name__in=["Recevio_Helpdesk", "SupportCenter"]).exists()

boxflow_required = user_passes_test(_can_boxflow)
helpdesk_required = user_passes_test(_can_helpdesk)


# --- AI Label Scan ---

def _call_ai_single(img_data, media_type):
    """OCR label image via Azure AI Vision, then extract sender/recipient with keyword parsing."""
    import requests as _req

    endpoint = os.environ.get("AZURE_VISION_ENDPOINT", "").rstrip("/")
    key = os.environ.get("AZURE_VISION_KEY", "")
    if not endpoint or not key:
        raise ValueError(
            "Ustaw AZURE_VISION_ENDPOINT i AZURE_VISION_KEY w konfiguracji Azure App Service. "
            "Utwórz zasób 'Computer Vision' w Azure Portal i skopiuj endpoint + klucz."
        )

    url = f"{endpoint}/computervision/imageanalysis:analyze?api-version=2024-02-01&features=read&language=pl"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": media_type}

    resp = _req.post(url, headers=headers, data=img_data, timeout=20)
    if resp.status_code == 401:
        raise ValueError("Błąd autoryzacji Azure Vision — sprawdź AZURE_VISION_KEY.")
    if resp.status_code == 404:
        raise ValueError("Nieprawidłowy AZURE_VISION_ENDPOINT — sprawdź adres zasobu Computer Vision.")
    resp.raise_for_status()

    # Collect all text lines from OCR result
    lines = []
    for block in resp.json().get("readResult", {}).get("blocks", []):
        for line in block.get("lines", []):
            text = line.get("text", "").strip()
            if text:
                lines.append(text)

    # Extract sender/recipient by looking for keyword labels or positional heuristics
    sender = ""
    recipient = ""
    SENDER_KW = {"nadawca", "sender", "from", "od", "nadawca:"}
    RECIPIENT_KW = {"odbiorca", "recipient", "to", "do", "adresat", "odbiorca:"}

    for i, line in enumerate(lines):
        lower = line.lower().rstrip(":")
        val_inline = line.split(":", 1)[1].strip() if ":" in line else ""
        if lower in SENDER_KW:
            sender = val_inline or (lines[i + 1] if i + 1 < len(lines) else "")
        elif lower in RECIPIENT_KW:
            recipient = val_inline or (lines[i + 1] if i + 1 < len(lines) else "")

    # Fallback: if no keywords found, first line = sender, second = recipient
    if not sender and not recipient and len(lines) >= 2:
        sender = lines[0]
        recipient = lines[1]

    return {"sender": sender.strip(), "recipient": recipient.strip()}


def _match_masked_name(raw_name, db_names):
    """
    Match a possibly-asterisked name (e.g. 'RAF*** ZAW***') against a list of DB names.
    Strategy (in order):
      1. Exact match after stripping asterisks (diacritic-insensitive)
      2. Standard fuzzy match (difflib) on the stripped name
      3. Token-prefix match: each visible fragment must be a prefix of some word in the candidate
    Returns the best matching DB name or None.
    """
    import re as _re
    import difflib as _difflib
    import unicodedata as _ud

    def _norm(s):
        return _ud.normalize("NFKD", s).encode("ascii", "ignore").decode().upper().strip()

    if not raw_name:
        return None

    # Strip asterisks → visible text
    stripped = _re.sub(r"\*+", " ", raw_name).strip()
    stripped = _re.sub(r"\s+", " ", stripped)
    norm_stripped = _norm(stripped)

    norm_db = {_norm(n): n for n in db_names}

    # 1. Exact (normalized)
    if norm_stripped in norm_db:
        return norm_db[norm_stripped]

    # 2. Standard fuzzy on stripped name
    close = _difflib.get_close_matches(norm_stripped, norm_db.keys(), n=1, cutoff=0.55)
    if close:
        return norm_db[close[0]]

    # 3. Token-prefix: every visible token (≥2 chars) must be a prefix of some word in candidate
    tokens = [t for t in _re.split(r"[\s*]+", raw_name.upper()) if len(t) >= 2]
    if tokens:
        best_name, best_score = None, 0
        for norm_candidate, orig_name in norm_db.items():
            cand_tokens = norm_candidate.split()
            score, matched_all = 0, True
            for tok in tokens:
                ntok = _norm(tok)
                if any(ct.startswith(ntok) for ct in cand_tokens):
                    score += len(tok)
                else:
                    matched_all = False
                    break
            if matched_all and score > best_score:
                best_score, best_name = score, orig_name
        if best_name:
            return best_name

    return None


@login_required
@boxflow_required
def boxflow_scan_label(request):
    """Step 1: Upload label image(s). Each image = one package; queued through confirm one by one."""
    if request.method == "POST":
        files = request.FILES.getlist("label_images") or (
            [request.FILES["label_image"]] if "label_image" in request.FILES else []
        )
        if not files:
            messages.error(request, "Please select at least one image.")
            return render(request, "boxflow/scan_label.html", {"form": LabelScanForm()})

        queue = []
        errors = []
        for f in files:
            try:
                result = _call_ai_single(f.read(), f.content_type or "image/jpeg")
                queue.append({
                    "sender": (result.get("sender") or "").strip(),
                    "recipient": (result.get("recipient") or "").strip(),
                })
            except Exception as e:
                errors.append(str(e))

        if errors:
            messages.warning(request, f"AI could not read {len(errors)} label(s): {errors[0]}")

        if not queue:
            return render(request, "boxflow/scan_label.html", {"form": LabelScanForm()})

        # Pop first into pack_prefill; store the rest as a queue for subsequent confirms
        request.session["pack_prefill"] = queue[0]
        if len(queue) > 1:
            request.session["pack_prefill_queue"] = queue[1:]
            messages.info(request, f"{len(queue)} packages scanned. Processing one by one.")
        return redirect("boxflow_add_confirm")

    return render(request, "boxflow/scan_label.html", {"form": LabelScanForm()})


# views.py
@login_required
@boxflow_required
@transaction.atomic
def boxflow_add_pack(request):
    """Step 2: Review/edit pre-filled form and save package."""
    prefill = request.session.pop("pack_prefill", None)

    if request.method == "POST":
        form = PackageForm(request.POST)
        if form.is_valid():
            # 1) Unikalny kod
            code = generate_package_code()
            while Package.objects.filter(code=code).exists():
                code = generate_package_code()

            # 2) Zapis paczki
            pkg = form.save(commit=False)
            if hasattr(pkg, "created_by"):
                pkg.created_by = request.user
            pkg.code = code
            pkg.status = Package.Status.IN_BOX
            pkg.save()
            form.save_m2m()

            # 3) Druk odbywa się po stronie klienta przez Zebra Browser Print

            # 4) E-mail do odbiorcy — fire-and-forget w tle, nie blokuje odpowiedzi
            if getattr(pkg.recipient, "email", None):
                _email = pkg.recipient.email
                _subject = f"Paczka w paczkomacie: {pkg.code}"
                _body = (
                    f"Cześć,\n\n"
                    f"Dla odbiorcy: {getattr(pkg.recipient, 'name', '')} zarejestrowano paczkę w paczkomacie.\n"
                    f"Kod: {pkg.code}\n"
                    f"Nadawca: {getattr(pkg.sender, 'name', '')}\n"
                    f"Dostarczono: {pkg.delivered_at:%d.%m.%Y %H:%M}\n\n"
                )
                Thread(
                    target=send_email_with_timeout,
                    args=(_email, _subject, _body),
                    kwargs={"timeout": 10},
                    daemon=True,
                ).start()

            # 5) If there are more packages in the batch queue, advance to the next one
            remaining = request.session.pop("pack_prefill_queue", [])
            messages.success(request, f"Package {pkg.code} has been added.")
            if remaining:
                request.session["pack_prefill"] = remaining[0]
                if len(remaining) > 1:
                    request.session["pack_prefill_queue"] = remaining[1:]
                messages.info(request, f"{len(remaining)} package(s) remaining in batch.")
                return redirect("boxflow_add_confirm")
            from django.urls import reverse
            detail_url = reverse("boxflow_detail", args=[pkg.pk])
            print_url = reverse("boxflow_print_label", args=[pkg.pk]) + f"?next={detail_url}"
            return redirect(print_url)

    else:
        # Pre-fill form with AI-extracted data
        initial = {}
        ai_sender_name = ""
        ai_recipient_name = ""
        if prefill:
            initial["delivered_at"] = timezone.now()

            # Try to match sender name to existing Sender (supports masked names like RAF*** ZAW***)
            sender_name = prefill.get("sender", "")
            if sender_name:
                from .models import Sender as _Sender
                all_senders = list(_Sender.objects.values_list("name", flat=True))
                matched_name = _match_masked_name(sender_name, all_senders)
                if matched_name:
                    initial["sender"] = _Sender.objects.get(name=matched_name)
                else:
                    # Pre-fill new_sender; form.save() will auto-create the Sender on submit
                    import re as _re
                    initial["new_sender"] = _re.sub(r"\*+", "", sender_name).strip()
                    ai_sender_name = sender_name

            # Try to match recipient name to existing Recipient (supports masked names)
            recipient_name = prefill.get("recipient", "")
            if recipient_name:
                from .models import Recipient as _Recipient
                all_recipients = list(_Recipient.objects.values_list("name", flat=True))
                matched_r_name = _match_masked_name(recipient_name, all_recipients)
                if matched_r_name:
                    initial["recipient"] = _Recipient.objects.get(name=matched_r_name)
                else:
                    ai_recipient_name = recipient_name

        form = PackageForm(initial=initial)

    return render(request, "boxflow/add_pack.html", {
        "form": form,
        "ai_recipient_name": ai_recipient_name if request.method == "GET" else "",
        "ai_sender_name": ai_sender_name if request.method == "GET" else "",
        "from_scan": prefill is not None,
    })



@login_required
@boxflow_required
def boxflow_pack_list(request):
    q = (request.GET.get("q") or "").strip()
    qs = (Package.objects
          .select_related("sender", "recipient")
          .order_by("-created_at"))
    if q:
        qs = qs.filter(
            Q(recipient__name__icontains=q) |
            Q(sender__name__icontains=q) |
            Q(code__icontains=q)
        )

    is_helpdesk = request.user.groups.filter(name="Recevio_Helpdesk").exists()

    return render(request, "boxflow/pack_list.html", {
        "packages": qs,
        "query": q,
        "is_helpdesk": is_helpdesk,   # ← tylko to dodajemy
    })


@login_required
@boxflow_required
def boxflow_inbox_status(request):
    qs = Package.objects.select_related("sender", "recipient").filter(
        status=Package.Status.IN_BOX
    ).order_by("-delivered_at")
    return render(request, "boxflow/inbox_status.html", {"packages": qs})

@login_required
@boxflow_required
def boxflow_pack_out(request):
    confirm = (request.method == "POST" and request.POST.get("action") == "confirm")
    form = ScanForm(request.POST or None, confirm=confirm)

    if request.method == "POST" and form.is_valid():
        code = form.cleaned_data["code"].strip()
        pkg = Package.objects.select_related("recipient").filter(code__iexact=code).first()
        if not pkg:
            messages.error(request, f"No package found with code: {code}")
            return redirect("boxflow_out")

        if pkg.status == Package.Status.ISSUED:
            who = pkg.collected_by.name if pkg.collected_by else (pkg.collected_by_name or "nieznana osoba")
            messages.info(request, f"Package {pkg.code} was already issued on {pkg.issued_at:%d.%m.%Y %H:%M}. Recipient: {who}.")
            return redirect("boxflow_out")

        # KROK 2: potwierdzenie
        if confirm:
            collected_by = form.cleaned_data.get("collected_by")
            collected_by_other = (form.cleaned_data.get("collected_by_other") or "").strip()

            issued_at = timezone.now()  # <— zamiast "now"
            pkg.status = Package.Status.ISSUED
            pkg.issued_at = issued_at
            pkg.issued_by = request.user
            pkg.collected_by = collected_by if collected_by else None
            pkg.collected_by_name = "" if collected_by else collected_by_other
            pkg.save(update_fields=[
                "status", "issued_at", "issued_by", "collected_by", "collected_by_name"
            ])

            who = collected_by.name if collected_by else collected_by_other

            # ✉️ e-mail — fire-and-forget w tle
            if getattr(pkg.recipient, "email", None):
                _email = pkg.recipient.email
                _subject = f"Paczka odebrana: {pkg.code}"
                _body = (
                    f"Cześć,\n\n"
                    f"Twoja paczka {pkg.code} została odebrana z paczkomatu.\n"
                    f"Odebrał(a): {who}\n"
                    f"Data i godzina odbioru: {issued_at:%d.%m.%Y %H:%M}\n"
                    f"Nadawca: {getattr(pkg.sender, 'name', '')}\n\n"
                )
                Thread(
                    target=send_email_with_timeout,
                    args=(_email, _subject, _body),
                    kwargs={"timeout": 10},
                    daemon=True,
                ).start()

            messages.success(request, f"Package {pkg.code} has been delivered. Recipient: {who}.")
            return redirect("boxflow_out")

        # KROK 1: poprawny skan → pokaż sekcję potwierdzenia w tej samej karcie
        if pkg.recipient_id:
            form.fields["collected_by"].initial = pkg.recipient_id

        return render(request, "boxflow/pack_out.html", {
            "form": form,
            "pkg": pkg,
            "show_confirm": True,   # ← UWAGA: nowa flaga
        })

    # GET lub BŁĄD w kroku 2 → pokazujemy formularz skanu; pola wyboru tylko gdy confirm było kliknięte
    context = {"form": form}
    if request.method == "POST" and confirm:
        # walidacja potwierdzenia nie przeszła – pokaż sekcję wyboru odbiorcy z błędami
        code = (request.POST.get("code") or "").strip()
        pkg = Package.objects.filter(code__iexact=code).first()
        context.update({"pkg": pkg, "show_confirm": True})
    return render(request, "boxflow/pack_out.html", context)


@login_required
@boxflow_required
def boxflow_pack_detail(request, pk):
    pkg = get_object_or_404(Package.objects.select_related("sender", "recipient"), pk=pk)
    is_helpdesk = request.user.groups.filter(name__in=["Recevio_Helpdesk", "SupportCenter"]).exists()
    return render(request, "boxflow/pack_detail.html", {"pkg": pkg, "is_helpdesk": is_helpdesk})


@login_required
@boxflow_required
def boxflow_reprint_label(request, pk):
    pkg = get_object_or_404(Package, pk=pk)

    # Jeśli zmieniłeś nazwę pliku, wskaż ją jawnie (pasuje do Twojego add_pack):
    template_path = os.path.join(
        settings.BASE_DIR, "GuestBook", "Print_templates", "BoxLabelTemplate.zpl"
    )

    try:
        zpl = render_box_label_from_file(
            code=pkg.code,
            sender=pkg.sender.name if getattr(pkg, "sender", None) else "",
            recipient=pkg.recipient.name if getattr(pkg, "recipient", None) else "",
            template_path=template_path,
        )

        _p_ip, _p_port = _get_printer_for_user(request.user)
        pr_status = run_with_timeout(send_zpl_to_printer, zpl,
                                     printer_ip=_p_ip, port=_p_port, seconds=5)

        if pr_status == "ok":
            messages.success(request, "The label has been reprinted.")
        elif pr_status == "timeout":
            messages.warning(request, "Label printing exceeded the time limit (5 seconds).")
        else:
            messages.error(request, "Label printing error.")

    except FileNotFoundError:
        messages.error(request, "The label template file was not found.")
    except Exception as e:
        messages.error(request, f"Printing error: {e}")

    return redirect("boxflow_detail", pk=pk)



# ============ HELPDESK / SUPPORTCENTER ============

@login_required
@helpdesk_required
def helpdesk_recipients(request):
    if request.method == "POST":
        action = request.POST.get("action") or "create"
        if action == "update":
            rid = request.POST.get("recipient_id")
            obj = get_object_or_404(Recipient, pk=rid)
            obj.name = (request.POST.get("name") or "").strip()
            email = (request.POST.get("email") or "").strip()
            obj.email = email or None
            obj.save(update_fields=["name", "email"])
            messages.success(request, "Recipient updated.")
            return redirect("helpdesk_recipients")
        else:
            form = RecipientForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Recipient added.")
                return redirect("helpdesk_recipients")
    else:
        form = RecipientForm()
    items = Recipient.objects.all().order_by("name")
    return render(request, "boxflow/helpdesk_recipients.html", {"form": form, "items": items})


@login_required
@helpdesk_required
def helpdesk_senders(request):
    if request.method == "POST":
        action = request.POST.get("action") or "create"
        if action == "update":
            sid = request.POST.get("sender_id")
            obj = get_object_or_404(Sender, pk=sid)
            obj.name = (request.POST.get("name") or "").strip()
            obj.save(update_fields=["name"])
            messages.success(request, "Sender updated.")
            return redirect("helpdesk_senders")
        else:
            form = SenderForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, "Sender added.")
                return redirect("helpdesk_senders")
    else:
        form = SenderForm()
    items = Sender.objects.all().order_by("name")
    return render(request, "boxflow/helpdesk_senders.html", {"form": form, "items": items})


@login_required
@helpdesk_required
def helpdesk_package_edit(request, pk):
    pkg = get_object_or_404(Package, pk=pk)
    if request.method == "POST":
        form = PackageEditForm(request.POST, instance=pkg)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.updated_by = request.user
            obj.save()
            messages.success(request, "Package changes saved.")
            return redirect("boxflow_detail", pk=pkg.pk)
    else:
        form = PackageEditForm(instance=pkg)
    return render(request, "boxflow/helpdesk_package_edit.html", {"form": form, "pkg": pkg})


# Import list z Excela (CSV/XLSX) – osobno dla nadawców i odbiorców
import io
import pandas as pd

@login_required
@helpdesk_required
def helpdesk_import_senders(request):
    if request.method == "POST" and request.FILES.get("file"):
        f = request.FILES["file"]
        df = pd.read_excel(f) if f.name.lower().endswith((".xls", ".xlsx")) else pd.read_csv(f)
        count = 0
        for name in df.iloc[:, 0].astype(str).str.strip():
            if name:
                Sender.objects.get_or_create(name=name)
                count += 1
        messages.success(request, f"{count} senders have been imported.")
        return redirect("helpdesk_senders")
    return render(request, "boxflow/helpdesk_import.html", {"title": "Import senders"})


@login_required
@helpdesk_required
def helpdesk_import_recipients(request):
    if request.method == "POST" and request.FILES.get("file"):
        f = request.FILES["file"]
        df = pd.read_excel(f) if f.name.lower().endswith((".xls", ".xlsx")) else pd.read_csv(f)
        # kolumna 1: name, kolumna 2 (opcjonalnie): email
        count = 0
        for _, row in df.iterrows():
            name = str(row.iloc[0]).strip()
            email = None
            if len(row) > 1 and pd.notna(row.iloc[1]):
                email = str(row.iloc[1]).strip()
            if name:
                obj, _ = Recipient.objects.get_or_create(name=name, defaults={"email": email})
                if email and not obj.email:
                    obj.email = email
                    obj.save(update_fields=["email"])
                count += 1
        messages.success(request, f"{count} recipients have been imported.")
        return redirect("helpdesk_recipients")
    return render(request, "boxflow/helpdesk_import.html", {"title": "Import recipients"})


@login_required
@user_passes_test(lambda u: u.is_authenticated and u.groups.filter(name="Recevio_Helpdesk").exists())
def boxflow_delete_pack(request, pk):
    pkg = get_object_or_404(Package, pk=pk)
    if request.method != 'POST':
        messages.error(request, "Deletion permitted only using the POST method.")
        return redirect('boxflow_list')
    code = pkg.code
    pkg.delete()
    messages.success(request, f"Package {code} has been delete.")
    return redirect('boxflow_list')


@require_POST
def kiosk_settings_save(request):
    """Save kiosk printer settings (password-protected)."""
    from .models import KioskSettings
    password = request.POST.get('password', '')
    if password != '0987':
        return JsonResponse({'ok': False, 'error': 'Wrong password'})
    ks = KioskSettings.get()
    ks.printer_address = request.POST.get('printer_address', ks.printer_address).strip()
    try:
        ks.printer_port = int(request.POST.get('printer_port', ks.printer_port))
    except ValueError:
        pass
    ks.save()
    return JsonResponse({'ok': True})


@login_required
@user_passes_test(lambda u: u.is_staff or u.groups.filter(name="Recevio_Helpdesk").exists())
def test_email_view(request):
    from django.core.mail import send_mail
    from django.conf import settings as django_settings

    config = {
        'EMAIL_HOST': getattr(django_settings, 'EMAIL_HOST', '—'),
        'EMAIL_PORT': getattr(django_settings, 'EMAIL_PORT', '—'),
        'EMAIL_USE_TLS': getattr(django_settings, 'EMAIL_USE_TLS', False),
        'EMAIL_USE_SSL': getattr(django_settings, 'EMAIL_USE_SSL', False),
        'EMAIL_HOST_USER': getattr(django_settings, 'EMAIL_HOST_USER', '—') or '(nie ustawiony)',
        'DEFAULT_FROM_EMAIL': getattr(django_settings, 'DEFAULT_FROM_EMAIL', '—'),
        'has_password': bool(getattr(django_settings, 'EMAIL_HOST_PASSWORD', '')),
    }

    result = None
    if request.method == 'POST':
        recipient = request.POST.get('recipient', '').strip()
        if recipient:
            try:
                send_mail(
                    subject='[Recevio] Test wiadomości e-mail',
                    message='To jest testowa wiadomość wysłana z systemu Recevio.\n\nJeśli ją otrzymałeś — konfiguracja SMTP działa poprawnie.',
                    from_email=django_settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[recipient],
                    fail_silently=False,
                )
                result = ('success', f'Wiadomość wysłana do: {recipient}')
            except Exception as e:
                result = ('error', str(e))
        else:
            result = ('error', 'Podaj adres e-mail odbiorcy.')

    return render(request, 'panel/test_email.html', {'config': config, 'result': result})


# ─── ZPL endpoints for Zebra Browser Print ───────────────────────────────────

@login_required
@boxflow_required
def boxflow_get_zpl(request, pk):
    """Returns ZPL for a package label as JSON — consumed by client-side Browser Print."""
    pkg = get_object_or_404(Package.objects.select_related("sender", "recipient"), pk=pk)
    template_path = os.path.join(settings.BASE_DIR, "GuestBook", "Print_templates", "BoxLabelTemplate.zpl")
    try:
        zpl = render_box_label_from_file(
            code=pkg.code,
            sender=pkg.sender.name if pkg.sender else "",
            recipient=pkg.recipient.name if pkg.recipient else "",
            template_path=template_path,
        )
        return JsonResponse({"zpl": zpl})
    except FileNotFoundError:
        return JsonResponse({"error": "Template file not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def visitor_get_zpl(request, pk):
    """Returns ZPL for a visitor badge as JSON — consumed by client-side Browser Print."""
    visitor = get_object_or_404(Visitor, pk=pk)
    template_path = os.path.join(settings.BASE_DIR, "GuestBook", "Print_templates", "template_zebra.zpl")
    try:
        with open(template_path, encoding="utf-8") as f:
            tpl = f.read()
        zpl = tpl.format(
            company=_company_display(visitor),
            first_name=visitor.first_name,
            last_name=visitor.last_name,
            visit_purpose=visitor.visit_purpose or "",
            supervisor=visitor.host.host_name if visitor.host else "",
        )
        return JsonResponse({"zpl": zpl})
    except FileNotFoundError:
        return JsonResponse({"error": "Template file not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@boxflow_required
def boxflow_print_label(request, pk):
    pkg = get_object_or_404(Package.objects.select_related("sender", "recipient"), pk=pk)
    next_url = request.GET.get("next", "")
    return render(request, "boxflow/print_label.html", {"pkg": pkg, "next_url": next_url})


@login_required
def visitor_print_badge(request, pk):
    visitor = get_object_or_404(Visitor, pk=pk)
    badge_css = None
    if visitor.production_area:
        badge_css = "red" if visitor.with_supervision else "green"
    return render(request, "panel/print_visitor_badge.html", {
        "visitor": visitor,
        "company": _company_display(visitor),
        "supervisor": visitor.host.host_name if visitor.host else "",
        "badge_css": badge_css,
    })


def public_pickup_view(request):
    """Public package pickup kiosk — no login required."""
    step = "scan"
    pkg = None
    error = None

    if request.method == "POST":
        action = request.POST.get("action", "scan")

        if action == "scan":
            code = (request.POST.get("code") or "").strip()
            if not code:
                error = "Zeskanuj lub wpisz kod paczki."
                step = "scan"
            else:
                pkg = Package.objects.select_related("sender", "recipient").filter(code__iexact=code).first()
                if not pkg:
                    error = f"Nie znaleziono paczki o kodzie: {code}"
                    step = "scan"
                elif pkg.status == Package.Status.ISSUED:
                    who = pkg.collected_by.name if pkg.collected_by else (pkg.collected_by_name or "nieznana osoba")
                    error = f"Paczka {pkg.code} została już odebrana przez: {who}."
                    step = "scan"
                else:
                    step = "confirm"

        elif action == "confirm":
            code = (request.POST.get("code") or "").strip()
            collected_by_name = (request.POST.get("collected_by_name") or "").strip()
            pkg = Package.objects.select_related("sender", "recipient").filter(code__iexact=code).first()

            if not pkg or pkg.status == Package.Status.ISSUED:
                error = "Nieprawidłowy stan paczki. Spróbuj ponownie."
                step = "scan"
            elif not collected_by_name:
                error = "Podaj imię i nazwisko odbierającego."
                step = "confirm"
            else:
                issued_at = timezone.now()
                pkg.status = Package.Status.ISSUED
                pkg.issued_at = issued_at
                pkg.collected_by = None
                pkg.collected_by_name = collected_by_name
                pkg.save(update_fields=["status", "issued_at", "collected_by", "collected_by_name"])

                if getattr(pkg.recipient, "email", None):
                    _subj = f"Paczka odebrana: {pkg.code}"
                    _body = (
                        f"Twoja paczka {pkg.code} została odebrana z paczkomatu.\n"
                        f"Odebrał(a): {collected_by_name}\n"
                        f"Data i godzina odbioru: {issued_at:%d.%m.%Y %H:%M}\n"
                        f"Nadawca: {getattr(pkg.sender, 'name', '')}\n"
                    )
                    _email = pkg.recipient.email
                    Thread(target=lambda: send_email_with_timeout(_email, _subj, _body, timeout=5), daemon=True).start()

                step = "done"

    packages = Package.objects.select_related("sender", "recipient").filter(
        status=Package.Status.IN_BOX
    ).order_by("-delivered_at")

    return render(request, "boxflow/public_pickup.html", {
        "step": step,
        "pkg": pkg,
        "error": error,
        "packages": packages,
    })
